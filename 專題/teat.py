import sys
import mysql.connector
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import random
from PyQt5 import uic, QtCore, QtGui
from PyQt5.QtGui import *
from PyQt5.QtWebEngineWidgets import *
from test1 import HtmlViewer
from chatbot_window import ChatbotWindow
import yfinance as yf
import hashlib, os
import mail
import requests
from lxml import html
import urllib.request
from lxml import etree
import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps)

class FetchStock:#观察清单数据获取
    def __init__(self, favorite_symbols):
        self.favorite_symbols = favorite_symbols

    def fetch_stock_data(self, symbol):
        market_codes = [".TW", ".TWO"]
        for market_code in market_codes:
            full_symbol = f"{symbol}{market_code}"
            stock = yf.Ticker(full_symbol)
            data = stock.history(period="5d")
            if not data.empty:
                close_price = data['Close'][-1]
                high_price = data['High'][-1]
                low_price = data['Low'][-1]
                open_price = data['Open'][-1]
                prev_close_price = data['Close'][-2] if len(data) > 1 else open_price
                price_change = close_price - prev_close_price
                price_change_percent = (price_change / prev_close_price) * 100
                return {
                    "symbol": symbol,
                    "market_code": market_code,
                    "open_price": f"{open_price:.2f}",
                    "close_price": f"{close_price:.2f}",
                    "high_price": f"{high_price:.2f}",
                    "low_price": f"{low_price:.2f}",
                    "prev_close_price": f"{prev_close_price:.2f}",
                    "price_change": f"{price_change:.2f}",
                    "price_change_percent": f"{price_change_percent:.2f}",
                    "is_favorite": symbol in self.favorite_symbols
                }
        return {"symbol": symbol, "error": "No data found"}


class SlidingPuzzleCaptcha(QWidget):
    def __init__(self, parent=None):
        super(SlidingPuzzleCaptcha, self).__init__(parent)
        self.initUI()
    
    def initUI(self):
        self.setFixedSize(300, 250)

        # 从指定文件夹中随机选择一张图片
        image_folder = "./test"  # 这里替换为你图片所在的文件夹路径
        self.bg_image = self.get_random_image(image_folder)

        self.block_size = 50
        self.block_x = random.randint(0, self.bg_image.width() - self.block_size)
        self.block_y = random.randint(0, self.bg_image.height() - self.block_size)
        self.image_with_hole = self.create_image_with_hole()

        self.image_label = QLabel(self)
        self.image_label.setPixmap(QPixmap.fromImage(self.image_with_hole))
        self.image_label.setFixedSize(self.bg_image.width(), self.bg_image.height())

        self.slider = QSlider(Qt.Horizontal, self)
        self.slider.setRange(0, self.bg_image.width() - self.block_size)
        self.slider.setTickPosition(QSlider.NoTicks)
        self.slider.setSingleStep(1)
        self.slider.valueChanged.connect(self.update_slider_position)
        self.slider.sliderReleased.connect(self.check_position)

        self.slider_block = QLabel(self)
        self.slider_block.setFixedSize(self.block_size, self.block_size)
        self.slider_block.setPixmap(QPixmap.fromImage(self.create_slider_block_image()))

        layout = QVBoxLayout(self)
        layout.addWidget(self.image_label)
        layout.addWidget(self.slider)
        layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(layout)

        QTimer.singleShot(100, lambda: self.update_slider_position(self.slider.value()))
    
    def get_random_image(self, folder):
        images = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f)) and f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif'))]
        if not images:
            raise FileNotFoundError("No images found in the specified directory.")
        random_image_path = os.path.join(folder, random.choice(images))
        return QImage(random_image_path)

    def create_image_with_hole(self):
        image = self.bg_image.copy()
        painter = QPainter(image)
        painter.setBrush(QBrush(Qt.white))
        painter.setPen(Qt.NoPen)
        painter.drawRect(self.block_x, self.block_y, self.block_size, self.block_size)
        painter.end()
        return image

    def create_slider_block_image(self):
        block_image = self.bg_image.copy(QRect(self.block_x, self.block_y, self.block_size, self.block_size))
        return block_image

    def update_slider_position(self, value):
        self.slider_block.move(value + self.image_label.x(), self.block_y + self.image_label.y())
    
    def check_position(self):
        # 断开信号以避免重复触发
        self.slider.blockSignals(True)
        
        slider_x = self.slider_block.x() - self.image_label.x()
        
        if abs(slider_x - self.block_x) < 5:
            QMessageBox.information(self, '驗證成功', '成功啦')
            self.slider.setDisabled(True)
            self.slider.blockSignals(False)  # 恢复信号
            self.parent().accept()  # 通知父窗口完成验证
        else:
            QMessageBox.warning(self, '驗證失敗', '拼錯啦再來一次吧')
            self.slider.setValue(0)
            self.slider.blockSignals(False)  # 恢复信号

class CaptchaDialog(QDialog):
    def __init__(self, parent=None):
        super(CaptchaDialog, self).__init__(parent)
        self.setWindowTitle("驗證")

        self.captcha_widget = SlidingPuzzleCaptcha(self)

        layout = QVBoxLayout(self)
        layout.addWidget(self.captcha_widget)
        self.setLayout(layout)
    def on_captcha_success(self):
        if self.captcha_widget.check_position():
            self.accept()
        else:
            self.captcha_widget.slider.setValue(0)

class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.db_connection = self.connect_to_db()
        self.verification_code = None
        self.initUI()
    def resizeEvent(self, event):
            super().resizeEvent(event)
            window_width = event.size().width()
            window_height = event.size().height()
            self.window_width = window_width
            self.window_height = window_height
            self.small_font_size = int(window_height/(525/9))
            self.big_font_size = int(window_height/(525/12))
            self.ex_big_font_size = int(window_height/(525/14))

            # 动态调整按钮的大小和位置
            self.frame_2.setGeometry(window_width/(962/20), window_height/(525/10), window_width/(962/441), window_height/(525/361))
            self.username.setGeometry(window_width/(962/120), window_height/(525/90), window_width/(962/211), window_height/(525/21))
            self.password.setGeometry(window_width/(962/120), window_height/(525/160), window_width/(962/211), window_height/(525/21))
            self.login_button.setGeometry(window_width/(962/110), window_height/(525/240), window_width/(962/241), window_height/(525/31))
            self.forget_password.setGeometry(window_width/(962/70), window_height/(525/300), window_width/(962/61), window_height/(525/20))
            self.quick_register.setGeometry(window_width/(962/340), window_height/(525/300), window_width/(962/61), window_height/(525/20))
            self.toggle_button.setGeometry(window_width/(962/340), window_height/(525/168), window_width/(962/41), window_height/(525/21))
            self.toggle_button.setIconSize(QSize(window_height/(525/41)/(41/14), window_height/(525/41)/(41/14))) 
            self.label_5.setScaledContents(True)

            self.stackedWidget.setGeometry(window_width/(962/240),window_height/(525/30),window_width/(962/461),window_height/(525/411),)
            self.username.setStyleSheet(f"""
                QLineEdit {{
                    background-color: transparent;
                    border: none; /* 移除边框 */
                    color: #717072; /* 设置文字颜色为灰色 */
                    border-bottom:1px solid #717072;
                    font: bold {self.big_font_size}px;
                }}
            """)    

            self.password.setStyleSheet(f"""
                QLineEdit {{
                    background-color: transparent;
                    border: none; /* 移除边框 */
                    color: #717072; /* 设置文字颜色为灰色 */
                    border-bottom:1px solid #717072;
                    font: bold {self.big_font_size}px;
                }}
            """)        
            self.quick_register.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent; /* 移除背景色 */
                    border: none; /* 移除边框 */
                    color: gray; /* 设置文字颜色，可以根据需要更改 */
                    font: bold {self.big_font_size}px; /* 设置文字样式和大小，可以根据需要更改 */
                }}
            """)            
            self.forget_password.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent; /* 移除背景色 */
                    border: none; /* 移除边框 */
                    color: gray; /* 设置文字颜色，可以根据需要更改 */
                    font: bold {self.big_font_size}px; /* 设置文字样式和大小，可以根据需要更改 */
                }}
            """)
            self.login_button.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px; 
                    font-size:{self.big_font_size}px;
                }}
            """)
            self.frame_7.setGeometry(window_width/(962/20), window_height/(525/10), window_width/(962/441), window_height/(525/361))
            self.find_email.setGeometry(window_width/(962/110), window_height/(525/120), window_width/(962/241), window_height/(525/21))
            self.tp_confirm_code.setGeometry(window_width/(962/110),window_height/(525/230), window_width/(962/241), window_height/(525/31))
            self.tp_confirm_code.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px; 
                    font-size:{self.big_font_size}px;
                }}
            """)

            self.frame_9.setGeometry(window_width/(962/20), window_height/(525/10), window_width/(962/441), window_height/(525/361))
            self.label.setGeometry(window_width/(962/110), window_height/(525/130), window_width/(962/61), window_height/(525/21))

            self.confirm_code.setGeometry(window_width/(962/180), window_height/(525/130), window_width/(962/161), window_height/(525/21))
            self.confirm_code_2.setGeometry(window_width/(962/110),window_height/(525/240), window_width/(962/241), window_height/(525/31))
            self.confirm_code_2.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px; 
                    font-size:{self.big_font_size}px;
                }}
            """)


            self.frame_13.setGeometry(window_width/(962/20), window_height/(525/10), window_width/(962/441), window_height/(525/361))
            self.label_3.setGeometry(window_width/(962/110), window_height/(525/70), window_width/(962/61), window_height/(525/21))
            self.label_4.setGeometry(window_width/(962/52), window_height/(525/170), window_width/(962/111), window_height/(525/21))
            self.new_password.setGeometry(window_width/(962/190), window_height/(525/70), window_width/(962/191), window_height/(525/20))
            self.confirm_code_6.setGeometry(window_width/(962/190), window_height/(525/170), window_width/(962/191), window_height/(525/20))

            self.re_new_code.setGeometry(window_width/(962/100),window_height/(525/240), window_width/(962/241), window_height/(525/31))
            self.re_new_code.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px; 
                    font-size:{self.big_font_size}px;
                }}
            """)
            self.label.setStyleSheet(f"""
                QLabel {{
                    color:white;                                    
                    font-size:{self.ex_big_font_size}px;    
                    font-weight: bold;
                }}
            """)


            self.label_3.setStyleSheet(f"""
                QLabel {{
                    color:white;font-size:{self.ex_big_font_size}px;    
                    font-weight: bold;
                }}
            """)
            self.label_4.setStyleSheet(f"""
                QLabel {{
                    color:white;font-size:{self.ex_big_font_size}px;    
                    font-weight: bold;
                }}
            """)



    def connect_to_db(self):
        try:
            connection = mysql.connector.connect(
                host='3.106.70.39',
                user='InvestUser',
                password='Zw.urVWv*gD@J5rT',
                database='aiinvest'
            )
            return connection
        except mysql.connector.Error as err:
            QMessageBox.critical(None, "Database Connection Error", f"Error: {err}")
            sys.exit()

    def close_connection(self):
        """關閉資料庫連線"""
        if self.db_connection and self.db_connection.is_connected():
            self.db_connection.close()
            print("已關閉 MySQL 連線")

    def initUI(self):
        uic.loadUi("./圖片/login.ui", self)

        window_width = 1080
        window_height = 991
        self.small_font_size = int(window_height/(525/9))
        self.big_font_size = int(window_height/(525/12))
        self.ex_big_font_size = int(window_height/(525/14))
        self.login_button.clicked.connect(self.handle_login)
        self.quick_register.clicked.connect(self.show_register)
        self.forget_password.clicked.connect(self.handle_forget_password)
        self.tp_confirm_code.clicked.connect(self.send_verification_code)
        self.confirm_code_2.clicked.connect(self.confirm_verification_code)
        self.re_new_code.clicked.connect(self.newpassword)
        self.toggle_button.setCheckable(True)
        self.toggle_button.setIcon(QIcon('./圖片/view.png'))
        self.toggle_button.setFixedSize(20, 20)
        self.toggle_button.setStyleSheet("QPushButton { border: none; }")
        self.password.setEchoMode(QLineEdit.Password)
        self.toggle_button.clicked.connect(self.toggle_password_visibility)
        self.frame_7.setGeometry(window_width/(962/20), window_height/(525/10), window_width/(962/441), window_height/(525/361))
        self.find_email.setGeometry(window_width/(962/110), window_height/(525/120), window_width/(962/241), window_height/(525/21))
        self.tp_confirm_code.setGeometry(window_width/(962/110),window_height/(525/230), window_width/(962/241), window_height/(525/31))
        self.tp_confirm_code.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px;	
                    font-size:{self.big_font_size}px;
                }}
            """)

        self.frame_9.setGeometry(window_width/(962/20), window_height/(525/10), window_width/(962/441), window_height/(525/361))
        self.label.setGeometry(window_width/(962/110), window_height/(525/130), window_width/(962/61), window_height/(525/21))

        self.confirm_code.setGeometry(window_width/(962/180), window_height/(525/130), window_width/(962/161), window_height/(525/21))
        self.confirm_code_2.setGeometry(window_width/(962/110),window_height/(525/240), window_width/(962/241), window_height/(525/31))
        self.confirm_code_2.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px;	
                    font-size:{self.big_font_size}px;
                }}
            """)


        self.frame_13.setGeometry(window_width/(962/20), window_height/(525/10), window_width/(962/441), window_height/(525/361))
        # self.label_2.setGeometry(window_width/(962/110), window_height/(525/60), window_width/(962/61), window_height/(525/21))
        self.label_3.setGeometry(window_width/(962/110), window_height/(525/120), window_width/(962/61), window_height/(525/21))
        self.label_4.setGeometry(window_width/(962/52), window_height/(525/170), window_width/(962/111), window_height/(525/21))
        # self.old_password.setGeometry(window_width/(962/190), window_height/(525/60), window_width/(962/191), window_height/(525/20))
        self.new_password.setGeometry(window_width/(962/190), window_height/(525/120), window_width/(962/191), window_height/(525/20))
        self.confirm_code_6.setGeometry(window_width/(962/190), window_height/(525/170), window_width/(962/191), window_height/(525/20))

        self.re_new_code.setGeometry(window_width/(962/100),window_height/(525/240), window_width/(962/241), window_height/(525/31))
        self.re_new_code.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px;	
                    font-size:{self.big_font_size}px;
                }}
            """)
        self.label.setStyleSheet(f"""
                QPushButton {{
                    color:white;
                    font-size:{self.ex_big_font_size}px;    
                    font-weight: bold;
                }}
            """)

        # self.label_2.setStyleSheet(f"""
        #         QPushButton {{
        #             color:white;font-size:{self.ex_big_font_size}px;    
        #             font-weight: bold;
        #         }}
        #     """)
        self.label_3.setStyleSheet(f"""
                QPushButton {{
                    color:white;font-size:{self.ex_big_font_size}px;    
                    font-weight: bold;
                }}
            """)
        self.label_4.setStyleSheet(f"""
                QPushButton {{
                    color:white;font-size:{self.ex_big_font_size}px;    
                    font-weight: bold;
                }}
            """)
    def toggle_password_visibility(self):
        if self.toggle_button.isChecked():
            self.password.setEchoMode(QLineEdit.Normal)
            self.toggle_button.setIcon(QIcon('./圖片/view.png'))
        else:
            self.password.setEchoMode(QLineEdit.Password)
            self.toggle_button.setIcon(QIcon('./圖片/hide.png'))

    def handle_login(self):
        username = self.username.text()
        password = self.password.text()

        hashed_password = hashlib.sha256(password.encode()).hexdigest()

        cursor = self.db_connection.cursor()
        query = "SELECT * FROM members WHERE name = %s AND password = %s"
        cursor.execute(query, (username, hashed_password))
        result = cursor.fetchone()
        cursor.close()
        if result:
            QMessageBox.information(None, "Login Success", "成功登錄")
            self.load_interface_ui(username)
        else:
            QMessageBox.warning(None, "Login Failed", "用戶名或者密碼錯誤")
    
    def show_register(self):
        self.register_window = RegisterWindow(self.db_connection,self.window_width,self.window_height)
        self.register_window.show()
        self.close()

    def load_interface_ui(self, username):
        self.interface_ui = IterfaceWindowLogined(username, self.db_connection,self.window_width,self.window_height)
        self.interface_ui.show()
        self.close()

    def handle_forget_password(self):
        self.stackedWidget.setCurrentIndex(1)
        current_page = self.stackedWidget.currentWidget()

        for child in current_page.findChildren(QWidget):
            print(f"控件名称: {child.objectName()}, 父级: {child.parent()}")

        for child in current_page.findChildren(QWidget):
            child.raise_()  # 提升控件

    def send_verification_code(self):
        self.input_email = self.find_email.text()
        try:
            # Create cursor and execute query
            cursor = self.db_connection.cursor()
            # Query email address
            query = "SELECT email FROM members WHERE email = %s"
            cursor.execute(query, (self.input_email,))
            result = cursor.fetchone()  # Use fetchone() to get a single result row
            if result:
                captcha_dialog = CaptchaDialog(self)
                if captcha_dialog.exec_() == QDialog.Accepted:
                    self.verification_code = mail.generate_secure_code()
                    mail.send_verifacation_email(self.input_email,self.verification_code)
                    self.stackedWidget.setCurrentIndex(2)
            else:
                QMessageBox.information(None, "fail", "此郵件尚未註冊")
                print("此郵件尚未註冊")
                self.stackedWidget.setCurrentIndex(0)
            # Commit changes if any
            self.db_connection.commit()
            
            # Close the cursor
            cursor.close()
        except mysql.connector.Error as err:
                QMessageBox.warning(None, "Registration Failed", f"Error: {err}")
                cursor.close()
                
    
    def confirm_verification_code(self):
        enter_code = self.confirm_code.text()
        if enter_code == self.verification_code:
            QMessageBox.information(None, "Verification Success", "驗證成功")
            # 進一步處理驗證成功的情況
            self.stackedWidget.setCurrentIndex(3)
        else:
            QMessageBox.warning(None, "Verification Failed", "驗證碼錯誤")
            # 處理驗證失敗的情況
            self.stackedWidget.setCurrentIndex(0)
 
    def newpassword(self):
        new_password = self.new_password.text()
        confirm_code = self.confirm_code_6.text()
        if new_password == confirm_code:
            try:
                cursor = self.db_connection.cursor()
                # 查詢郵件是否存在於資料庫中
                check_query = "SELECT * FROM members WHERE email = %s"
                cursor.execute(check_query, (self.input_email,))
                # 獲取查詢結果
                result = cursor.fetchone()
                if result:
                    # 對密碼進行雜湊處理
                    hashed_password = hashlib.sha256(new_password.encode()).hexdigest()
                    # 郵件存在，更新密碼
                    update_query = "UPDATE members SET password = %s WHERE email = %s"
                    cursor.execute(update_query, (hashed_password, self.input_email))
                    # 提交變更至資料庫
                    self.db_connection.commit()
                    QMessageBox.information(None, "Success", "變更密碼成功")
                    self.stackedWidget.setCurrentIndex(0)
                    cursor.close()
            except Exception as e:
                # 處理可能的例外情況
                QMessageBox.information(None, "Error", f"發生錯誤：{str(e)}")
        else:
            # 密碼不一致
            QMessageBox.information(None, "Error", "輸入密碼不一致")
            self.stackedWidget.setCurrentIndex(3)

    def back_to_login(self):
        self.stackedWidget.setCurrentIndex(0)

class RegisterWindow(QWidget):#注册画面
    def __init__(self, db_connection,window_width,window_height):
        super().__init__()
        self.db_connection = db_connection
        self.initUI()
        self.resize(window_width, window_height)
    def validate_password(self):
        if not self.password.hasAcceptableInput():
            # 设置错误时的样式
            self.password.setStyleSheet("""
                QLineEdit {
                    background-color: transparent;
                    border: none;
                    color: red;
                    border-bottom: 1px solid #717072;
                }
            """)
        else:
            self.password.setStyleSheet("""
                QLineEdit {
                    background-color: transparent;
                    border: none;
                    color: #717072;
                    border-bottom: 1px solid #717072;
                }
            """)
    def validate_email(self):
        if not self.email.hasAcceptableInput():
            # 设置错误时的样式
            self.email.setStyleSheet("""
                QLineEdit {
                    background-color: transparent;
                    border: none;
                    color: red;
                    border-bottom: 1px solid #717072;
                }
            """)
        else:
            self.email.setStyleSheet("""
                QLineEdit {
                    background-color: transparent;
                    border: none;
                    color: #717072;
                    border-bottom: 1px solid #717072;
                }
            """)
    def resizeEvent(self, event):
            super().resizeEvent(event)
            window_width = event.size().width()
            window_height = event.size().height()
            self.window_width = window_width
            self.window_height = window_height
            self.small_font_size = int(window_height/(525/9))
            self.big_font_size = int(window_height/(525/12))
            self.ex_big_font_size = int(window_height/(525/14))

            # 动态调整按钮的大小和位置
            self.frame.setGeometry(0, 0, window_width/(962/979), window_height/(525/525))

            self.frame_2.setGeometry(window_width/(962/330), window_height/(525/110), window_width/(962/301), window_height/(525/291))
            self.username.setGeometry(window_width/(962/60), window_height/(525/70), window_width/(962/161), window_height/(525/20))
            self.password.setGeometry(window_width/(962/60), window_height/(525/130), window_width/(962/161), window_height/(525/20))
            self.register_button.setGeometry(window_width/(962/60), window_height/(525/220), window_width/(962/191), window_height/(525/20))
            self.confirm_password.setGeometry(window_width/(962/60), window_height/(525/160), window_width/(962/161), window_height/(525/20))
            self.email.setGeometry(window_width/(962/60), window_height/(525/100), window_width/(962/161), window_height/(525/20))
            self.register_button.setStyleSheet(f"""
                QPushButton {{
                    background-color: red; /* 设置背景颜色为红色 */
                    color: white; /* 设置文字颜色为白色 */
                    border: none; /* 可选：移除边框 */
                    font-weight: 300;
                    padding: 6px; /* 设置内边距，使按钮更大一些 */
                    border-radius: 5px;	
                    font-size:{self.big_font_size}px;
                }}
            """)
            self.label_2.setScaledContents(True)
    def initUI(self):
        uic.loadUi("./圖片/register.ui", self)
        
        # 设置邮箱验证的正则表达式
        email_regex = QRegularExpression(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")
        self.email_validator = QRegularExpressionValidator(email_regex, self.email)
        self.email.setValidator(self.email_validator)
        
        # 设置密码验证的正则表达式（8-16个字符，只能是数字或字母）
        password_regex = QRegularExpression(r"^[a-zA-Z0-9]{8,16}$")
        self.password_validator = QRegularExpressionValidator(password_regex, self.password)
        self.password.setValidator(self.password_validator)
        
        # 连接textChanged信号到自定义槽函数
        self.email.textChanged.connect(self.validate_email)
        self.password.textChanged.connect(self.validate_password)
        
        self.register_button.clicked.connect(self.handle_register)

    def handle_register(self):
        username = self.username.text()
        email = self.email.text()
        password = self.password.text()
        confirm_password = self.confirm_password.text()

        # 检查邮箱格式是否有效
        if not self.email.hasAcceptableInput():
            QMessageBox.warning(self, "Registration Failed", "非法邮箱")
            return

        # 检查密码格式是否有效
        if not self.password.hasAcceptableInput():
            QMessageBox.warning(self, "Registration Failed", "密码必须是8-16个字符，只能包含数字或字母")
            return

        if confirm_password != password:
            QMessageBox.warning(self, "Registration Failed", "密码不一致，请重新输入")
            return

        cursor = self.db_connection.cursor()
        check_query = "SELECT * FROM members WHERE name = %s"
        cursor.execute(check_query, (username,))
        result = cursor.fetchone()
        cursor.close()
        if result:
            QMessageBox.warning(self, "Registration Failed", "已经有人取过这个名字了")
        else:
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            query = "INSERT INTO members (name, email, password) VALUES (%s, %s, %s)"
            try:
                cursor = self.db_connection.cursor()
                cursor.execute(query, (username, email, hashed_password))
                self.db_connection.commit()
                cursor.close()
                QMessageBox.information(self, "Registration Success", "成功註冊!")
                self.show_interface(self.username.text())
            except mysql.connector.Error as err:
                QMessageBox.warning(self, "Registration Failed", f"Error: {err}")
                cursor.close()
    def show_interface(self,username):
        self.interface_window = IterfaceWindowLogined(username, self.db_connection,self.window_width,self.window_height)
        self.interface_window.show()
        self.close()

class NumericTableWidgetItem(QTableWidgetItem):  # 自定义表格单元格类(用來排序數字用)
    def __lt__(self, other):
        # 嘗試將值轉換為浮點數進行比較
        try:
            return float(self.text()) < float(other.text())
        except ValueError:
            # 如果無法轉換為數字，按字母順序排序
            return self.text() < other.text()

class IterfaceWindowLogined(QWidget):#登录后画面
    def __init__(self, username, db_connection,window_width,window_height):
        super().__init__()
        self.db_connection = db_connection
        self.username = username
        self.data = []
        self.initUI()
        self.fade_out_animation_connected = False
        self.load_stock_data()
        self.start_timer()
        self.resize(window_width,window_height)
        self.overlay = None

    def connect_to_db(self):
        try:
            connection = mysql.connector.connect(
                host='3.106.70.39',
                user='InvestUser',
                password='Zw.urVWv*gD@J5rT',
                database='aiinvest'
            )
            return connection
        except mysql.connector.Error as err:
            QMessageBox.critical(None, "Database Connection Error", f"Error: {err}")
            sys.exit()

    def close_connection(self):
        """關閉資料庫連線"""
        if self.db_connection and self.db_connection.is_connected():
            self.db_connection.close()
            print("已關閉 MySQL 連線")

    def resizeEvent(self, event):
        # 获取当前窗口大小
        window_width = event.size().width()
        window_height = event.size().height()
        self.window_width = window_width
        self.window_height = window_height
        # 动态调整按钮的大小和位置
        self.frame.setGeometry(0, 0, window_width, window_height/(525/41))
        self.pushButton_4.setGeometry(window_width/(962/840),window_height/(525/10),window_width/(962/61),window_height/(525/21),)
        self.pushButton_4.setIconSize(QSize(window_height/(525/41)/(41/20), window_height/(525/41)/(41/20))) 
        self.pushButton_6.setGeometry(window_width/(962/930),0,window_width/(962/21),window_height/(525/20),)
        self.pushButton_6.setIconSize(QSize(window_height/(525/41)/(41/10), window_height/(525/41)/(41/10))) 

        self.lineEdit.setGeometry(window_width/(962/880),window_height/(525/10),window_width/(962/81),window_height/(525/41)/(41/20),)
        self.lineEdit_font_size = int(window_height/(525/14))
        self.lineEdit.setStyleSheet(f"""
            font-size: {self.lineEdit_font_size}px;
            background: transparent;
            border: 0px solid transparent;
            padding: 0px;
            margin: 0px;
            color: black;
        """)
        self.frame_4.setGeometry(0,window_height/(525/40) , window_width/(962/41), window_height/(525/491))
        self.pushButton_7.setGeometry(window_width/(962/5),window_height/(525/20) , window_width/(962/31), window_height/(525/31))
        self.pushButton_7.setIconSize(QSize(window_height/(525/41)/(41/20), window_height/(525/41)/(41/20))) 

        self.show_chatbot.setGeometry(window_width/(962/5),window_height/(525/350) , window_width/(962/31), window_height/(525/31))
        self.show_chatbot.setIconSize(QSize(window_height/(525/41)/(41/25), window_height/(525/41)/(41/25))) 


        self.pushButton_3.setGeometry(window_width/(962/5),window_height/(525/60) , window_width/(962/31), window_height/(525/41))
        self.pushButton_3.setIconSize(QSize(window_height/(525/41)/(41/20), window_height/(525/41)/(41/20))) 

        self.pushButton_5.setGeometry(window_width/(962/5),window_height/(525/112) , window_width/(962/31), window_height/(525/31))
        self.pushButton_5.setIconSize(QSize(window_height/(525/41)/(41/20), window_height/(525/41)/(41/20))) 

        self.public_bt.setGeometry(window_width/(962/5),window_height/(525/150) , window_width/(962/31), window_height/(525/31))
        self.public_bt.setIconSize(QSize(window_height/(525/41)/(41/25), window_height/(525/41)/(41/25))) 

        self.toolButton.setGeometry(window_width/(962/14),window_height/(525/3) , window_width/(962/211), window_height/(525/31))
        self.toolButton.setIconSize(QSize(window_height/(525/41)/(41/250), window_height/(525/41)/(41/150))) 
        self.stackedWidget.setGeometry(window_width/(962/40),window_height/(525/40) , window_width/(962/921), window_height/(525/481))
        self.viewer.setGeometry(0,0 , window_width/(962/640), window_height/(525/480))
        self.tableWidget.setGeometry(0,window_height/(525/10) , window_width/(962/911), window_height/(525/261))

        self.tableWidget_mylist.setGeometry(10,window_height/(525/40) , window_width/(962/931), window_height/(525/401))
        self.widget.setGeometry(-1,-1 , window_width/(962/921), window_height/(525/41))
        self.pushButton_8.setGeometry(0,0 , window_width/(962/121), window_height/(525/31))
        self.pushButton_8.setIconSize(QSize(window_height/(525/41)/(41/12), window_height/(525/41)/(41/12))) 


        self.tableWidget_mylist_2.setGeometry(0,0 , window_width/(962/921), window_height/(525/451))
        self.table_widget.setGeometry(window_width/(962/710),window_height/(525/40) , window_width/(962/256), window_height/(525/281))
        self.tabWidget.setGeometry(0,window_height/(525/50) , window_width/(962/931), window_height/(525/441))
        self.widget_edit_st.setGeometry(0,window_height/(525/200) , window_width/(962/911), window_height/(525/171))
        self.comboBox.setGeometry(window_width/(962/135),window_height/(525/90) , window_width/(962/125), window_height/(525/22))
        self.lineEdit_2.setGeometry(window_width/(962/135),window_height/(525/55) , window_width/(962/125), window_height/(525/22))
        self.lineEdit_3.setGeometry(window_width/(962/135),window_height/(525/125) , window_width/(962/125), window_height/(525/22))
        self.lineEdit_4.setGeometry(window_width/(962/135),window_height/(525/160) , window_width/(962/125), window_height/(525/22))
        self.dateEdit.setGeometry(window_width/(962/520),window_height/(525/55) , window_width/(962/125), window_height/(525/22))
        self.dateEdit_2.setGeometry(window_width/(962/520),window_height/(525/100) , window_width/(962/125), window_height/(525/22))
        self.checkBox.setGeometry(window_width/(962/400),window_height/(525/160) , window_width/(962/131), window_height/(525/21))
        self.checkBox_2.setGeometry(window_width/(962/400),window_height/(525/180) , window_width/(962/161), window_height/(525/21))
        self.pushButton_9.setGeometry(window_width/(962/680),window_height/(525/160) , window_width/(962/90), window_height/(525/30))
        self.pushButton_9_font_size = int(window_height/(525/9))
        self.new_pushButton_9_font_size = int(window_height/(525/12))
        self.pushButton_9.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.new_pushButton_9_font_size}px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #1FB6F6, stop: 1 #497BF0
                );
                color: white;
                border: 1px solid #2980b9;
                border-radius: 7px;
            }}

            QPushButton:pressed {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #1FB6F6, stop: 1 #497BF0
                );
            }}
        """)


        self.label.setGeometry(window_width/(962/30),window_height/(525/55) , window_width/(962/51), window_height/(525/21))
        self.label_3.setGeometry(window_width/(962/30),window_height/(525/90) , window_width/(962/51), window_height/(525/21))
        self.label_4.setGeometry(window_width/(962/400),window_height/(525/55) , window_width/(962/51), window_height/(525/21))
        self.label_5.setGeometry(window_width/(962/400),window_height/(525/90) , window_width/(962/51), window_height/(525/21))
        self.label_7.setGeometry(window_width/(962/390),window_height/(525/10) , window_width/(962/81), window_height/(525/21))
        self.label_6.setGeometry(window_width/(962/20),window_height/(525/10) , window_width/(962/81), window_height/(525/21))
        self.label_8.setGeometry(window_width/(962/30),window_height/(525/125) , window_width/(962/91), window_height/(525/21))
        self.label_9.setGeometry(window_width/(962/30),window_height/(525/160) , window_width/(962/91), window_height/(525/21))
        self.label_10.setGeometry(window_width/(962/390),window_height/(525/125) , window_width/(962/81), window_height/(525/21))
        self.label_11.setGeometry(window_width/(962/400),window_height/(525/200) , window_width/(962/261), window_height/(525/21))
        self.small_label_font_size = int(window_height/(525/12))
        self.big_label_font_size = int(window_height/(525/14))
        self.pushButton_8_font_size = int(window_height/(525/12))
        self.ex_small_label_font_size = int(window_height/(525/8))


        self.pushButton_8.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #1FB6F6, stop: 1 #497BF0
                );
                color: white;
                border: 1px solid #2980b9;
                border-radius: 7px;
                font-size: {self.pushButton_8_font_size}px;
            }}
            QPushButton:pressed {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #1FB6F6, stop: 1 #497BF0
                );
            }}


        """)
        self.label.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.label_3.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.label_4.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.label_5.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.label_6.setStyleSheet(f"""
            font-size: {self.big_label_font_size}px;
            font-weight: bold;
        """)
        self.label_7.setStyleSheet(f"""
            font-size: {self.big_label_font_size}px;
            font-weight: bold;
        """)
        self.label_8.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.label_9.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.label_10.setStyleSheet(f"""
            font-size: {self.big_label_font_size}px;
            font-weight: bold;
        """)
        self.label_11.setStyleSheet(f"""
            font-size: {self.ex_small_label_font_size}px;
            font-weight: bold;
            color: red;
        """)
        self.checkBox.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.checkBox_2.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.tableWidget.setColumnWidth(0, self.window_width/(962/130))
        self.tableWidget.setColumnWidth(1, self.window_width/(962/130))
        self.tableWidget.setColumnWidth(2, self.window_width/(962/80))
        self.tableWidget.setColumnWidth(3, self.window_width/(962/80))
        self.tableWidget.setColumnWidth(4, self.window_width/(962/80))
        self.tableWidget.setColumnWidth(5, self.window_width/(962/80))
        self.tableWidget.setColumnWidth(6, self.window_width/(962/80))

        self.tableWidget_2.setColumnWidth(0, self.window_width/(962/80))
        self.tableWidget_2.setColumnWidth(1, self.window_width/(962/80))
        self.tableWidget_2.setColumnWidth(2, self.window_width/(962/80))
        self.tableWidget_2.setColumnWidth(3, self.window_width/(962/80))
        self.tableWidget_2.setColumnWidth(4, self.window_width/(962/130))
        self.tableWidget_2.setColumnWidth(5, self.window_width/(962/130))
        self.tableWidget_2.setColumnWidth(6, self.window_width/(962/80))
        self.tableWidget_2.setColumnWidth(7, self.window_width/(962/130))

        self.textBrowser.setFixedWidth(self.window_width)

        self.tableWidget_3.setColumnWidth(0, self.window_width/(962/80))
        self.tableWidget_3.setColumnWidth(1, self.window_width/(962/80))
        self.tableWidget_3.setColumnWidth(2, self.window_width/(962/80))
        self.tableWidget_3.setColumnWidth(3, self.window_width/(962/80))
        self.tableWidget_3.setColumnWidth(4, self.window_width/(962/80))
        self.tableWidget_3.setColumnWidth(5, self.window_width/(962/80))
        self.tableWidget_3.setColumnWidth(6, self.window_width/(962/130))
        self.tableWidget_3.setColumnWidth(7, self.window_width/(962/130))
        self.tableWidget_3.setColumnWidth(8, self.window_width/(962/80))
        self.tableWidget_3.setColumnWidth(9, self.window_width/(962/130))

        self.tableWidget_4.setColumnWidth(0, self.window_width/(962/80))
        self.tableWidget_4.setColumnWidth(1, self.window_width/(962/80))
        self.tableWidget_4.setColumnWidth(2, self.window_width/(962/80))
        self.tableWidget_4.setColumnWidth(3, self.window_width/(962/80))
        self.tableWidget_4.setColumnWidth(4, self.window_width/(962/80))
        self.tableWidget_4.setColumnWidth(5, self.window_width/(962/130))
        self.tableWidget_4.setColumnWidth(6, self.window_width/(962/130))
        self.tableWidget_4.setColumnWidth(7, self.window_width/(962/80))
        self.tableWidget_4.setColumnWidth(8, self.window_width/(962/130))

        self.table_widget.setColumnWidth(0, self.window_width/(962/10))
        self.table_widget.setColumnWidth(1, self.window_width/(962/60))
        self.table_widget.setColumnWidth(2, self.window_width/(962/60))
        self.table_widget.setColumnWidth(3, self.window_width/(962/50))
        self.table_widget.setColumnWidth(4, self.window_width/(962/50))

        self.tableWidget_mylist.setColumnWidth(0, self.window_width/(962/70))
        self.tableWidget_mylist.setColumnWidth(1, self.window_width/(962/600))
        self.tableWidget_mylist.setColumnWidth(2, self.window_width/(962/35))
        self.tableWidget_mylist.setColumnWidth(3, self.window_width/(962/35))
        self.tableWidget_mylist.setColumnWidth(4, self.window_width/(962/35))
        self.tableWidget_mylist.setRowHeight(0, self.window_height/(525/30))
        self.tableWidget_mylist.setRowHeight(1, self.window_height/(525/30))
        self.tableWidget_mylist.setRowHeight(2, self.window_height/(525/30))
        self.tableWidget_mylist.setRowHeight(3, self.window_height/(525/30))
        self.tableWidget_mylist.setRowHeight(4, self.window_height/(525/30))


        self.table_widget.setRowHeight(0, self.window_height/(525/30))
        self.table_widget.setRowHeight(1, self.window_height/(525/30))
        self.table_widget.setRowHeight(2, self.window_height/(525/30))
        self.table_widget.setRowHeight(3, self.window_height/(525/30))
        self.table_widget.setRowHeight(4, self.window_height/(525/30))



        self.tableWidget_mylist_2.setColumnWidth(0, self.window_width/(962/70))
        self.tableWidget_mylist_2.setColumnWidth(1, self.window_width/(962/600))
        self.tableWidget_mylist_2.setColumnWidth(2, self.window_width/(962/35))
        self.tableWidget_mylist_2.setRowHeight(0, self.window_height/(525/30))
        self.tableWidget_mylist_2.setRowHeight(1, self.window_height/(525/30))
        self.tableWidget_mylist_2.setRowHeight(2, self.window_height/(525/30))

        self.fill_example_data()
        self.set_tableWidget_mylist_font()
        self.set_table_widget_font()
        self.reset_tableWidget_font()
        self.reset_tableWidget_2_font()
        # self.reset_textBrowser_font()
        self.reset_tableWidget_3_font()
        self.reset_tableWidget_4_font()
        self.reset_tableWidget_mylist_2_font()
        geometry = self.viewer.geometry()
        # print(geometry)
        # print(window_width,window_height)

        super().resizeEvent(event)


    def initUI(self):#一些界面UI,UX设定基本不动
        uic.loadUi("./圖片/interface_logined.ui", self)
        self.window_height = 525

        screen = QApplication.primaryScreen()
        dpi = screen.logicalDotsPerInch()
        print(f"Screen DPI: {dpi}")
        scale_factor = int(dpi / 96)
        self.button_edit = QPushButton()
        self.button_execute = QPushButton()
        self.button_delete = QPushButton()

        # 设置窗口无边框
        # 设置圆角效果
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.dateEdit.setStyleSheet("""
                QCalendarWidget QToolButton#qt_calendar_prevmonth{                    
                    qproperty-iconSize: 14px;   /* 自定义箭头图标大小 */
                    qproperty-icon: url(./圖片/back.png);  /* 资源文件路径 */

                }
                QCalendarWidget QToolButton#qt_calendar_nextmonth {
                    qproperty-iconSize: 14px;   /* 自定义箭头图标大小 */
                    qproperty-icon: url(./圖片/next.png);  /* 资源文件路径 */
                }
            """)
        self.dateEdit_2.setStyleSheet("""
                QCalendarWidget QToolButton#qt_calendar_prevmonth{                    
                    qproperty-iconSize: 14px;   /* 自定义箭头图标大小 */
                    qproperty-icon: url(./圖片/back.png);  /* 资源文件路径 */

                }
                QCalendarWidget QToolButton#qt_calendar_nextmonth {
                    qproperty-iconSize: 14px;   /* 自定义箭头图标大小 */
                    qproperty-icon: url(./圖片/next.png);  /* 资源文件路径 */
                }
            """)



        self.pushButton_6.clicked.connect(self.close)
        self.pushButton_7.setToolTip("主頁")
        self.pushButton_3.setToolTip("我的策略")
        self.pushButton_5.setToolTip("觀察清單")
        self.public_bt.setToolTip("公共策略")
        self.lineEdit.setReadOnly(True)
        self.lineEdit.setText(self.username)
        self.parameter_data = ['0056.TW','KDCross','2014-07-26','2024-07-26','code','10000','0.002'] #參數默認設定 
        self.parameter50_data_default = [['2330', '2317', '2454', '2308', '2881', '2382', '2303', '2882', '2891', '3711',
                                        '2412', '2886', '2884', '1216', '2357', '2885', '2892', '2327', '3034', '2890',
                                        '5880', '2345', '3231', '2880', '3008', '2883', '2002', '2379', '4938', '2207',
                                        '1303', '2887', '1101', '2603', '2301', '3037', '1301', '5871', '3017', '3045',
                                        '2912', '4904', '6446', '2395', '6669', '3661', '5876', '1326', '1590', '6505'],
                                        'KDCross','2014-07-26','2024-07-26','code','10000','0.002'] #參數默認設定 
        self.parameter50_name = ['台積電', '鴻海', '聯發科', '台達電', '富邦金', '廣達', '聯電', '國泰金', '中信金', '日月光投控',
                                 '中華電', '兆豐金', '玉山金', '統一', '華碩', '元大金', '第一金', '國巨', '聯詠', '永豐金',
                                 '合庫金', '智邦', '緯創', '華南金', '大立光', '開發金', '中鋼', '瑞昱', '和碩', '和泰車', '南亞',
                                 '台新金', '台泥', '長榮', '光寶科', '欣興', '台塑', '中租-KY', '奇鋐', '台灣大', '統一超',
                                 '遠傳', '藥華藥', '研華', '緯穎', '世芯-KY', '上海商銀', '台化', '亞德客-KY', '台塑化']
        page_2 = self.stackedWidget.widget(1)
        page_2_layout = QVBoxLayout(page_2)
        self.splitter = QSplitter(Qt.Vertical)
        self.splitter.setStyleSheet("QSplitter::handle { background-color: transparent; }")
        self.viewer = HtmlViewer('./HTML/white.html')
        self.splitter.addWidget(self.viewer)
        self.tabWidget = QTabWidget()

        tab1 = QWidget()
        self.tab1_layout = QVBoxLayout(tab1)
        self.tab1_layout.addWidget(self.widget_edit_st)
        self.tabWidget.addTab(tab1, "策略參數調整")

        tab2 = QWidget()
        self.tab2_layout = QVBoxLayout(tab2)
        self.tab2_layout.addWidget(self.tableWidget)
        self.tabWidget.addTab(tab2, "交易明細")

        tab3 = QWidget()
        self.tab3_layout = QVBoxLayout(tab3)
        self.tab3_layout.addWidget(self.tableWidget_2)
        self.tab3_layout.addWidget(self.textBrowser)
        self.tabWidget.addTab(tab3, "交易成果")


        tab4 = QWidget()
        self.tab4_layout = QVBoxLayout(tab4)
        self.tab4_layout.addWidget(self.tableWidget_3)
        self.tabWidget.addTab(tab4, "0050成分股之交易成果")
        self.index_of_tab_4 = self.tabWidget.indexOf(tab4)
        if self.checkBox.isChecked():
            # 顯示 tab_4
            self.tabWidget.setTabVisible(self.index_of_tab_4, True)
        else:
            # 隱藏 tab_4
            self.tabWidget.setTabVisible(self.index_of_tab_4, False)

        tab5 = QWidget()
        self.tab5_layout = QVBoxLayout(tab5)
        self.tab5_layout.addWidget(self.tableWidget_4)
        self.tabWidget.addTab(tab5, "觀察清單成分股之交易成果")
        self.my_parameter_data = [[],'KDCross','2014-07-26','2024-07-26','code','10000','0.002']
        if (self.check_target() == None):
            self.checkBox_2.setChecked(False)
            self.checkBox_2.setEnabled(False)
        else:
            self.checkBox_2.setEnabled(True)
            self.my_parameter_data[0] = []
            for id in self.check_target():
                self.my_parameter_data[0].append(id)
        self.index_of_tab_5 = self.tabWidget.indexOf(tab5)
        if self.checkBox_2.isChecked():
            # 顯示 tab_5
            self.tabWidget.setTabVisible(self.index_of_tab_5, True)
        else:
            # 隱藏 tab_5
            self.tabWidget.setTabVisible(self.index_of_tab_5, False)

        self.tabWidget.setMinimumHeight(45) 
        self.splitter.addWidget(self.tabWidget)
        self.splitter.setStretchFactor(0, 3)
        self.splitter.setStretchFactor(1, 4)
        self.splitter.splitterMoved.connect(self.handle_splitter_moved)

        page_2_layout.addWidget(self.splitter)

        page_2.setLayout(page_2_layout)
        self.pushButton_7.clicked.connect(self.go_to_page_1)
        self.pushButton_3.clicked.connect(self.go_to_page_2)
        temp = self.find_my_strategy()
        # print(temp)
        self.comboBox.clear()
        if temp:  # 如果有结果
            for strategy in temp:
                self.comboBox.addItem(strategy[0])  # 将每个策略添加到 combobox

        self.pushButton_9.clicked.connect(self.st_temp)
        self.public_bt.clicked.connect(self.load_public_list)
        self.pushButton_7.enterEvent = self.create_hover_event(self.pushButton_7, 1.2)
        self.pushButton_7.leaveEvent = self.create_leave_event(self.pushButton_7)
        self.pushButton_3.enterEvent = self.create_hover_event(self.pushButton_3, 1.2)
        self.pushButton_3.leaveEvent = self.create_leave_event(self.pushButton_3)
        self.pushButton_5.enterEvent = self.create_hover_event(self.pushButton_5, 1.2)
        self.pushButton_5.leaveEvent = self.create_leave_event(self.pushButton_5)
        self.public_bt.enterEvent = self.create_hover_event(self.public_bt, 1.2)
        self.public_bt.leaveEvent = self.create_leave_event(self.public_bt)
        self.show_chatbot.enterEvent = self.create_hover_event(self.show_chatbot, 1.2)
        self.show_chatbot.leaveEvent = self.create_leave_event(self.show_chatbot)
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(5)
        self.table_widget.setHorizontalHeaderItem(1, QTableWidgetItem('股票代號'))
        self.table_widget.setHorizontalHeaderItem(2, QTableWidgetItem('最新價'))
        self.table_widget.setHorizontalHeaderItem(3, QTableWidgetItem('漲跌'))
        self.table_widget.setHorizontalHeaderItem(4, QTableWidgetItem('漲跌%'))
        self.icon_add = QIcon('./圖片/add.png')
        header_item = QTableWidgetItem()
        header_item.setIcon(self.icon_add)
        header_item.setText("   ")
        header_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table_widget.setHorizontalHeaderItem(0, header_item)
        self.icon1 = QIcon('./圖片/delete.png')
        for row in range(self.table_widget.rowCount()):
            if self.table_widget.item(row, 1):
                item_with_icon = QTableWidgetItem()
                item_with_icon.setIcon(self.icon1)
                item_with_icon.setText("")
                self.table_widget.setItem(row, 0, item_with_icon)
        self.table_widget.cellClicked.connect(self.on_cell_clicked)
        self.table_widget.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.setVisible(False)
        self.tableWidget.setColumnWidth(0, 30)
        self.tableWidget.setColumnWidth(1, 30)
        self.tableWidget_2.setColumnWidth(0, 30)
        self.tableWidget_2.setColumnWidth(1, 30)
        self.tableWidget_3.setColumnWidth(0, 30)
        self.tableWidget_3.setColumnWidth(1, 30)
        self.tableWidget_4.setColumnWidth(0, 30)
        self.tableWidget_4.setColumnWidth(1, 30)
        self.table_widget.setColumnWidth(0, 30)
        self.table_widget.setColumnWidth(1, 30)
        self.table_widget.setColumnWidth(2, 30)
        self.table_widget.setColumnWidth(3, 30)
        self.table_widget.setColumnWidth(4, 30)
        self.tableWidget_mylist.setColumnWidth(0, 30)
        self.tableWidget_mylist.setColumnWidth(1, 30)
        self.tableWidget_mylist.setColumnWidth(2, 30)
        self.tableWidget_mylist.setColumnWidth(3, 30)
        self.tableWidget_mylist.setColumnWidth(4, 30)

        self.pushButton_8.clicked.connect(self.show_increase)
        self.fill_example_data()

        self.table_widget.setShowGrid(False)

        self.pushButton_3.clicked.connect(self.fill_example_data)
        self.pushButton_5.clicked.connect(self.toggle_table_widget_visibility)
        self.pushButton_4.clicked.connect(self.show_member_info)


        # self.load_trades_data_defualt()

        # 获取UI中的控件

        self.chatbot_window = ChatbotWindow()
        self.chatbot_window.setWindowFlags(Qt.FramelessWindowHint)


        self.show_chatbot.clicked.connect(self.toggle_chatbot)
        # 初始化聊天窗口为隐藏状态

        # 连接按钮点击事件到方法
    def toggle_chatbot(self):
        username = self.username
        user_id = self.get_member_id()
        # 使用者資訊
        botpress_url = f'https://927f-2001-e10-6840-107-7150-8c4-b902-d1fb.ngrok-free.app/botpress?username={username}&user_id={user_id}'
        # 打開該網址
        webbrowser.open(botpress_url)

    def handle_splitter_moved(self, pos, index):
        total_height = self.splitter.height()

        # 获取各个 widget 的高度
        viewer_height = self.viewer.height()
        tabWidget_height = self.tabWidget.height()

        # 设置 tabWidget 的最小高度阈值
        min_tabWidget_height = 200

        # 如果 tabWidget 的高度小于或等于最小值，限制 splitter
        if tabWidget_height <= min_tabWidget_height:
            # 计算新的大小分配，保持 viewer 的高度不变，只调整 tabWidget 的高度
            self.splitter.setSizes([viewer_height, total_height - viewer_height])

            # 如果调整后 tabWidget 仍然小于最小值，再次修正
            if self.tabWidget.height() < min_tabWidget_height:
                self.splitter.setSizes([viewer_height, min_tabWidget_height])
    def get_start_date(self):
        self.dateEdit_start.date().toString('yyyy-MM-dd')

    def get_end_date(self):
        return self.dateEdit_end.date().toString('yyyy-MM-dd')
    def st_temp(self):
        if self.lineEdit_2.text() == '' or self.comboBox.currentText() == '' or self.lineEdit_3.text() == '' or self.lineEdit_4.text() == '':
            QMessageBox.information(None, "Warning", "請輸入完整資料")
        elif self.dateEdit.date().toString('yyyy-MM-dd') == self.dateEdit_2.date().toString('yyyy-MM-dd') or self.dateEdit.date().toString('yyyy-MM-dd') > self.dateEdit_2.date().toString('yyyy-MM-dd'):
            QMessageBox.information(None, "Warning", "起始日期須小於結束日期")
        elif not (-0.10 <= float(self.lineEdit_4.text()) < 0.10):
            QMessageBox.information(None, "Warning", "最大交易手續費必須在 -0.1 到 0.1 之間")
        else:
            self.parameter_data[0] = self.lineEdit_2.text()
            self.parameter_data[0] += '.TW'
            self.parameter_data[1] = self.comboBox.currentText()
            # print(self.parameter_data[1])

            self.parameter_data[2] = self.dateEdit.date().toString('yyyy-MM-dd')
            self.parameter_data[3] = self.dateEdit_2.date().toString('yyyy-MM-dd')
            self.parameter_data[5] = self.lineEdit_3.text()
            self.parameter_data[6] = self.lineEdit_4.text()

            if self.checkBox.isChecked():
                self.parameter50_data = self.parameter50_data_default.copy()
                for i in range(len(self.parameter50_data[0])):
                    self.parameter50_data[0][i] = self.parameter50_data[0][i].replace('.TW','')
                    self.parameter50_data[0][i] += '.TW'
                self.parameter50_data[1] = self.comboBox.currentText()
                self.parameter50_data[2] = self.dateEdit.date().toString('yyyy-MM-dd')
                self.parameter50_data[3] = self.dateEdit_2.date().toString('yyyy-MM-dd')
                self.parameter50_data[5] = self.lineEdit_3.text()
                self.parameter50_data[6] = self.lineEdit_4.text()

            if self.checkBox_2.isChecked():
                for i in range(len(self.my_parameter_data[0])):
                    self.my_parameter_data[0][i] = self.my_parameter_data[0][i].replace('.TW','')
                    self.my_parameter_data[0][i] += '.TW'
                self.my_parameter_data[1] = self.comboBox.currentText()
                self.my_parameter_data[2] = self.dateEdit.date().toString('yyyy-MM-dd')
                self.my_parameter_data[3] = self.dateEdit_2.date().toString('yyyy-MM-dd')
                self.my_parameter_data[5] = self.lineEdit_3.text()
                self.my_parameter_data[6] = self.lineEdit_4.text()
            # print(self.my_parameter_data)
            self.load_trades_data()

    def check_target(self):
        cursor = self.db_connection.cursor(buffered=True)
        member_id = self.get_member_id()
        query = "SELECT target_id FROM target WHERE member_id = %s"
        cursor.execute(query, (member_id, ))
        result = cursor.fetchall()
        result = [row[0] for row in result]
        if result == []:
            cursor.close()
            return None
        cursor.close()
        return result

    def find_my_strategy(self):
        self.member_id = self.get_member_id()
        cursor = None
        try:
            cursor = self.db_connection.cursor()
            query = "SELECT strategy FROM code WHERE member_id = %s AND public = 0"
            cursor.execute(query, (self.member_id,))
            results = cursor.fetchall()  # 获取所有查询结果
            return results  # 返回结果列表

        except mysql.connector.Error as err:
            QMessageBox.warning(self, "Database Error", f"Error: {err}")
            return []

        finally:
            if cursor:
                cursor.close()  # 确保游标在操作完成后关闭
    def paintEvent(self, event):#圆角动画
        path = QPainter(self)
        path.setRenderHint(QPainter.Antialiasing)
        path.setBrush(QBrush(QColor(255, 255, 255)))
        path.setPen(Qt.NoPen)
        path.drawRoundedRect(self.rect(), 15, 15)
    def load_public_list(self):  # 加载公共策略画面
        cursor = self.db_connection.cursor()
        try:
            m_id = self.get_member_id()
            
            # 查询所有 public = 1 的策略
            query_public = """
            SELECT strategy, description  
            FROM code 
            WHERE public = 1
            """
            cursor.execute(query_public)
            public_strategies = cursor.fetchall()

            # 查询当前用户已经拥有的相同策略 (public = 0)
            owned_query = """
            SELECT strategy FROM code WHERE member_id = %s AND public = 0
            """
            cursor.execute(owned_query, (m_id,))
            owned_strategies = {row[0] for row in cursor.fetchall()}

        finally:
            cursor.close()  # 确保游标被关闭


        self.tableWidget_mylist_2.setRowCount(len(public_strategies))
        self.tableWidget_mylist_2.setColumnCount(3)


        for row_index, row_data in enumerate(public_strategies):
            strategy_name = row_data[0]
            for col_index, cell_data in enumerate(row_data):
                self.tableWidget_mylist_2.setItem(row_index, col_index, QTableWidgetItem(cell_data))

            # 根据用户是否已经拥有该策略 (public = 0) 来设置按钮
            button_increase = QPushButton()

            if strategy_name in owned_strategies:
                button_increase.setIcon(QIcon('./圖片/forbidden.png'))

                button_increase.setEnabled(False)
            else:
                button_increase.setIcon(QIcon('./圖片/downloads.png'))
                button_increase.clicked.connect(lambda _, row_index=row_index, button=button_increase: self.handle_strategy_click(row_index, button))

            self.tableWidget_mylist_2.setCellWidget(row_index, 2, button_increase)
            
            button_increase.setStyleSheet("""
            QPushButton {
                border: none;
                padding: 5px 10px;
            }

            """)

        self.reset_tableWidget_mylist_2_font()
        self.stackedWidget.setCurrentIndex(3)

    def handle_strategy_click(self, row_index, button):
        # 设置为已拥有的图标
        button.setIcon(QIcon('./圖片/forbidden.png'))
        button.setEnabled(False)
        # 处理新增策略逻辑
        self.add_strategy(row_index)
        temp = self.find_my_strategy()
        self.comboBox.clear()
        if temp:  # 如果有结果
            for strategy in temp:
                self.comboBox.addItem(strategy[0])  # 将每个策略添加到 combobox
    def add_strategy(self, row_index):
        strategy = self.tableWidget_mylist_2.item(row_index, 0).text()
        description = self.tableWidget_mylist_2.item(row_index, 1).text()
        file_name = str(self.get_member_id()) + "_" + strategy

        try:
            cursor = self.db_connection.cursor()
            query = """
            INSERT INTO code (strategy, description, member_id, file_name) 
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
            description = VALUES(description), 
            member_id = VALUES(member_id)
            """
            cursor.execute(query, (strategy, description, self.get_member_id(), file_name))
            self.db_connection.commit()
            import test3
            test3.test3_main(self.get_member_id(), strategy)

        except mysql.connector.Error as err:
            QMessageBox.warning(self, "Database Error", f"Error: {err}")

        finally:
            cursor.close()  # 确保游标被关闭
        
        # 刷新页面
        self.fill_example_data()
        self.load_public_list()
            

    # def fill_example_data(self):  # 填充個人策略
    #     # # 每次查詢前重新建立資料庫連線
    #     self.close_connection()
    #     self.db_connection = self.connect_to_db()  # 確保這個方法存在以建立新連線
    #     cursor = self.db_connection.cursor()
    #     try:
    #         m_id = self.get_member_id()

    #         # 查询符合当前 m_id 的 strategy 和 description
    #         query = """
    #         SELECT strategy, description 
    #         FROM code 
    #         WHERE member_id = %s AND public = 0
    #         """
    #         cursor.execute(query, (m_id,))
    #         # self.db_connection.commit()
    #         example_data = cursor.fetchall()
    #         print(example_data)

    #     finally:
    #         cursor.close()  # 确保游标被关闭

    #     self.tableWidget_mylist.setRowCount(len(example_data))
    #     self.tableWidget_mylist.setColumnCount(5)
   

    #     for row_index, row_data in enumerate(example_data):
    #         for col_index, cell_data in enumerate(row_data):
    #             self.tableWidget_mylist.setItem(row_index, col_index, QTableWidgetItem(cell_data))
    #         self.add_button_to_table(row_index, 2, 3 ,4)
    #     self.comboBox.clear()
    #     temp = self.find_my_strategy()
    #     if temp:  # 如果有结果
    #         for strategy in temp:
    #             self.comboBox.addItem(strategy[0])  # 将每个策略添加到 combobox
    #     self.set_tableWidget_mylist_font()
    #     self.reset_tableWidget_mylist_bt()
    #     print("我的策略頁面重製")
    def fill_example_data(self):
        # 每次呼叫時重新建立資料庫連線
        conn = mysql.connector.connect(
            host='3.106.70.39',
            user='InvestUser',
            password='Zw.urVWv*gD@J5rT',
            database='aiinvest'
        )
        cursor = conn.cursor()

        try:
            m_id = self.get_member_id()
            query = """
                SELECT strategy, description 
                FROM code 
                WHERE member_id = %s AND public = 0
            """
            cursor.execute(query, (m_id,))
            example_data = cursor.fetchall()
        finally:
            cursor.close()
            conn.close()

        # 更新表格的行數
        self.tableWidget_mylist.setRowCount(len(example_data))
        self.tableWidget_mylist.setColumnCount(5)

        for row_index, row_data in enumerate(example_data):
            for col_index, cell_data in enumerate(row_data):
                self.tableWidget_mylist.setItem(row_index, col_index, QTableWidgetItem(str(cell_data)))

            # 假設有 add_button_to_table 方法，將按鈕加入到指定位置
            self.add_button_to_table(row_index, 2, 3, 4)

        # 更新 ComboBox
        self.comboBox.clear()
        temp = self.find_my_strategy()
        if temp:
            for strategy in temp:
                self.comboBox.addItem(strategy[0])

        self.set_tableWidget_mylist_font()
        self.reset_tableWidget_mylist_bt()

    def set_table_widget_font(self):
        
        font = QFont("Arial", 8*(self.window_height/525))  # 设置全局字体大小 150%7 125%12 约1.7倍
        row_count = self.table_widget.rowCount()
        col_count = self.table_widget.columnCount()
        header_font = int(self.window_height/(525/12))

        # 遍历所有单元格并设置字体
        for row in range(row_count):
            for col in range(col_count):
                item = self.table_widget.item(row, col)
                if item:
                    item.setFont(font)
        self.table_widget.setStyleSheet(f"""
            QTableWidget {{
                border: black;  /* 隐藏整个表格的边框 */
                gridline-color: transparent;  /* 隐藏单元格的网格线 */
                background-color: rgba(255, 255, 255, 150)

            }}

            QHeaderView::section {{
                font-size: {header_font}px;

            }}

            QPushButton {{
                border: none;  /* 隐藏按钮的边框 */
            }}
            QTableWidget::item:selected {{
            background-color: transparent;   /* 移除选中时的背景颜色 */
            color:black;
                }}
            QTableWidget::item:focus {{
                background-color: transparent;   /* 移除焦点时的背景颜色 */
            }}
        """)

    def set_tableWidget_mylist_font(self):
        font = QFont("Arial", 12*(int(self.window_height/525)))  # 设置全局字体大小 150%7 125%12 约1.7倍
        row_count = self.tableWidget_mylist.rowCount()
        col_count = self.tableWidget_mylist.columnCount()
        header_font = int(self.window_height/(525/12))

        # 遍历所有单元格并设置字体
        for row in range(row_count):
            for col in range(col_count):
                item = self.tableWidget_mylist.item(row, col)
                if item:
                    item.setFont(font)
        self.tableWidget_mylist.setStyleSheet(f"""
            QTableWidget {{
                border: none;  /* 隐藏整个表格的边框 */
                gridline-color: transparent;  /* 隐藏单元格的网格线 */
            }}

            QHeaderView::section {{
                font-size: {header_font}px;
                border: none;  /* 隐藏列标签的边框 */
            }}

            QPushButton {{
                border: none;  /* 隐藏按钮的边框 */
            }}
            QTableWidget::item:selected {{
            background-color: transparent;   /* 移除选中时的背景颜色 */
            color:black;
                }}
            QTableWidget::item:focus {{
                background-color: transparent;   /* 移除焦点时的背景颜色 */
            }}
        """)

    def reset_tableWidget_font(self):
        font = QFont("Arial", 7*(self.window_height/525))  # 设置全局字体大小 150%7 125%12 约1.7倍
        row_count = self.tableWidget.rowCount()
        col_count = self.tableWidget.columnCount()
        header_font = int(self.window_height/(525/12))
        # 遍历所有单元格并设置字体
        for row in range(row_count):
            for col in range(col_count):
                item = self.tableWidget.item(row, col)
                if item:
                    item.setFont(font)
        self.tableWidget.setStyleSheet(f"""
            QTableWidget {{
                border: none;  /* 隐藏整个表格的边框 */
                gridline-color: transparent;  /* 隐藏单元格的网格线 */
            }}

            QHeaderView::section {{
                font-size: {header_font}px;
                border: none;  /* 隐藏列标签的边框 */
            }}

            QPushButton {{
                border: none;  /* 隐藏按钮的边框 */
            }}
            QTableWidget::item:selected {{
            background-color: transparent;   /* 移除选中时的背景颜色 */
            color:black;
                }}
            QTableWidget::item:focus {{
                background-color: transparent;   /* 移除焦点时的背景颜色 */
            }}
        """)
    def reset_tableWidget_2_font(self):
        font = QFont("Arial", 7*(self.window_height/525))  # 设置全局字体大小 150%7 125%12 约1.7倍
        row_count = self.tableWidget_2.rowCount()
        col_count = self.tableWidget_2.columnCount()
        header_font = int(self.window_height/(525/12))
        # 遍历所有单元格并设置字体
        for row in range(row_count):
            for col in range(col_count):
                item = self.tableWidget_2.item(row, col)
                if item:
                    item.setFont(font)
        self.tableWidget_2.setStyleSheet(f"""
            QTableWidget {{
                border: none;  /* 隐藏整个表格的边框 */
                gridline-color: transparent;  /* 隐藏单元格的网格线 */
            }}

            QHeaderView::section {{
                font-size: {header_font}px;
                border: none;  /* 隐藏列标签的边框 */
            }}

            QPushButton {{
                border: none;  /* 隐藏按钮的边框 */
            }}
            QTableWidget::item:selected {{
            background-color: transparent;   /* 移除选中时的背景颜色 */
            color:black;
                }}
            QTableWidget::item:focus {{
                background-color: transparent;   /* 移除焦点时的背景颜色 */
            }}
        """)
    def reset_tableWidget_3_font(self):
        font = QFont("Arial", 7*(self.window_height/525))  # 设置全局字体大小 150%7 125%12 约1.7倍
        row_count = self.tableWidget_3.rowCount()
        col_count = self.tableWidget_3.columnCount()
        header_font = int(self.window_height/(525/12))
        # 遍历所有单元格并设置字体
        for row in range(row_count):
            for col in range(col_count):
                item = self.tableWidget_3.item(row, col)
                if item:
                    item.setFont(font)
        self.tableWidget_3.setStyleSheet(f"""
            QTableWidget {{
                border: none;  /* 隐藏整个表格的边框 */
                gridline-color: transparent;  /* 隐藏单元格的网格线 */
            }}

            QHeaderView::section {{
                font-size: {header_font}px;
                border: none;  /* 隐藏列标签的边框 */
            }}

            QPushButton {{
                border: none;  /* 隐藏按钮的边框 */
            }}
            QTableWidget::item:selected {{
            background-color: transparent;   /* 移除选中时的背景颜色 */
            color:black;
                }}
            QTableWidget::item:focus {{
                background-color: transparent;   /* 移除焦点时的背景颜色 */
            }}
        """)
    def reset_tableWidget_4_font(self):
        font = QFont("Arial", 7*(self.window_height/525))  # 设置全局字体大小 150%7 125%12 约1.7倍
        row_count = self.tableWidget_4.rowCount()
        col_count = self.tableWidget_4.columnCount()
        header_font = int(self.window_height/(525/12))
        # 遍历所有单元格并设置字体
        for row in range(row_count):
            for col in range(col_count):
                item = self.tableWidget_4.item(row, col)
                if item:
                    item.setFont(font)
        self.tableWidget_4.setStyleSheet(f"""
            QTableWidget {{
                border: none;  /* 隐藏整个表格的边框 */
                gridline-color: transparent;  /* 隐藏单元格的网格线 */
            }}

            QHeaderView::section {{
                font-size: {header_font}px;
                border: none;  /* 隐藏列标签的边框 */
            }}

            QPushButton {{
                border: none;  /* 隐藏按钮的边框 */
            }}
            QTableWidget::item:selected {{
            background-color: transparent;   /* 移除选中时的背景颜色 */
            color:black;
                }}
            QTableWidget::item:focus {{
                background-color: transparent;   /* 移除焦点时的背景颜色 */
            }}
        """)
    def reset_tableWidget_mylist_2_font(self):
        font = QFont("Arial", 7*(self.window_height/525))  # 设置全局字体大小 150%7 125%12 约1.7倍
        row_count = self.tableWidget_mylist_2.rowCount()
        col_count = self.tableWidget_mylist_2.columnCount()
        header_font = int(self.window_height/(525/12))
        # 遍历所有单元格并设置字体
        for row in range(row_count):
            for col in range(col_count):
                item = self.tableWidget_mylist_2.item(row, col)
                if item:
                    item.setFont(font)
        self.tableWidget_mylist_2.setStyleSheet(f"""
            QTableWidget {{
                border: none;  /* 隐藏整个表格的边框 */
                gridline-color: transparent;  /* 隐藏单元格的网格线 */
            }}

            QHeaderView::section {{
                font-size: {header_font}px;
                border: none;  /* 隐藏列标签的边框 */
            }}

            QPushButton {{
                border: none;  /* 隐藏按钮的边框 */
            }}
            QTableWidget::item:selected {{
            background-color: transparent;   /* 移除选中时的背景颜色 */
            color:black;
                }}
            QTableWidget::item:focus {{
                background-color: transparent;   /* 移除焦点时的背景颜色 */
            }}
        """)
    def reset_tableWidget_mylist_bt(self):
        self.button_edit_font_size = int(self.window_height/(525/15))
        # print(self.button_edit_font_size)
        self.button_edit.setStyleSheet(f"""
            font-size: {self.button_edit_font_size}px;
        """)
        self.button_edit.setIconSize(QSize(int(self.window_height/(525/41)/(41/15)), int(self.window_height/(525/41)/(41/15)))) 
        self.button_execute.setStyleSheet(f"""
            font-size: {self.button_edit_font_size}px;
        """)
        self.button_delete.setIconSize(QSize(int(self.window_height/(525/41)/(41/15)), int(self.window_height/(525/41)/(41/15)))) 
        self.button_delete.setStyleSheet(f"""
            font-size: {self.button_edit_font_size}px;
        """)
        self.button_execute.setIconSize(QSize(QSize(int(self.window_height/(525/41)/(41/15)), int(self.window_height/(525/41)/(41/15)))))

    def add_button_to_table(self, row, column_edit, column_execute, column_delete):
        self.button_edit = QPushButton()
        self.button_execute = QPushButton()    
        self.button_delete = QPushButton()

        # 设置删除按钮的图标和样式
        edit_icon = QIcon('./圖片/editing.png')  
        self.button_edit.setIcon(edit_icon)

        execute_icon = QIcon('./圖片/output.png')  
        self.button_execute.setIcon(execute_icon)
        delete_icon = QIcon('./圖片/delete.png')  
        self.button_delete.setIcon(delete_icon)
        self.button_delete.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
        """)

        self.button_edit.setStyleSheet("""
                QPushButton {
                background-color: transparent;
                border: none;
            }
        """)

        self.button_execute.setStyleSheet("""
                QPushButton {
                background-color: transparent;
                border: none;
            }
        """)
        self.button_edit.clicked.connect(lambda: self.on_edit_button_clicked(row))
        self.button_execute.clicked.connect(lambda: self.on_excute_button_clicked(row))
        self.button_delete.clicked.connect(lambda: self.on_delete_button_clicked(row))
        self.reset_tableWidget_mylist_bt()


        container_edit = QWidget()
        container_execute = QWidget()
        container_delete = QWidget()

        layout_edit = QVBoxLayout()
        layout_execute = QVBoxLayout()
        layout_delete = QVBoxLayout()

        layout_edit.addWidget(self.button_edit)
        layout_execute.addWidget(self.button_execute)
        layout_delete.addWidget(self.button_delete)

        layout_edit.setContentsMargins(0, 0, 0, 0)
        layout_execute.setContentsMargins(0, 0, 0, 0)
        layout_delete.setContentsMargins(0, 0, 0, 0)

        container_edit.setLayout(layout_edit)
        container_execute.setLayout(layout_execute)
        container_delete.setLayout(layout_delete)

        self.tableWidget_mylist.setCellWidget(row, column_edit, container_edit)
        self.tableWidget_mylist.setCellWidget(row, column_execute, container_execute)
        self.tableWidget_mylist.setCellWidget(row, column_delete, container_delete)
    def on_delete_button_clicked(self, row):
        try:
            # 获取选中的策略名称（假设策略名称在index 0的列）
            strategy = self.tableWidget_mylist.item(row, 0).text()
            member_id = self.get_member_id()

            # 显示确认对话框
            reply = QMessageBox.question(
                self, 
                "確認刪除", 
                f"你確定要刪除 '{strategy}'嗎?永遠都回不來了喔。",
                QMessageBox.Yes | QMessageBox.No, 
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                # 创建数据库游标
                db_cursor = self.db_connection.cursor()

                # 删除数据库中的记录
                delete_query = """
                DELETE FROM code 
                WHERE member_id = %s AND strategy = %s
                """
                db_cursor.execute(delete_query, (member_id, strategy))

                # 提交事务
                self.db_connection.commit()

                # 从表格中移除这一行
                self.tableWidget_mylist.removeRow(row)

                # 关闭游标
                db_cursor.close()

                # 提示删除成功
                QMessageBox.information(self, "删除成功", f"策略 '{strategy}' 已成功删除。")
            else:
                # 如果用户选择了 "No"，取消删除操作
                QMessageBox.information(self, "取消操作", "策略删除已取消。")

        except mysql.connector.Error as err:
            QMessageBox.warning(self, "Database Error", f"Error: {err}")

        finally:
            # 确保 db_cursor 被成功初始化后才关闭
            if 'db_cursor' in locals() and db_cursor is not None:
                db_cursor.close()
    def on_edit_button_clicked(self,row):
        print(f"编辑按钮在行 {row} 被点击")
       
        self.show_edit(row) #show出edit画面
    def on_excute_button_clicked(self, row):

        print(f"執行按钮在行 {row} 被点击")
        self.parameter_data[1] =self.tableWidget_mylist.item(row, 0).text()
        self.load_trades_data()
        # print(self.parameter_data)


    def update_html(self):
        if self.viewer is not None:
            self.splitter.widget(0).deleteLater()  # 移除并删除当前 viewer

    def load_trades_data_defualt(self): #加载交易明细画面（默認）
        font = QFont("Arial", 7*(self.window_height/525))
        # import test1
        # rslt, trades = test1.test1_main(self.parameter_data,self.member_id, 0) #總之這種的就是呼叫test1 parameter_data是給test1的參數ctrl+f可以找到在哪
        # self.tableWidget.setRowCount(len(trades))
        # for row_index, row_data in enumerate(trades.itertuples(index=False)):
        #     formatted_data = [
        #         row_data.EntryTime.date(),
        #         row_data.ExitTime.date(),
        #         round(row_data.EntryPrice),
        #         round(row_data.ExitPrice),
        #         round(row_data.Size),
        #         round(row_data.PnL),
        #         round(row_data.ReturnPct,2)
        #     ]
        #     for col_index, cell_data in enumerate(formatted_data): #展示不用動
        #         item = QTableWidgetItem(str(cell_data))
        #         item.setFont(font)
        #         item.setTextAlignment(Qt.AlignCenter)
        #         self.tableWidget.setItem(row_index, col_index, item) 
        # self.tableWidget_2.setRowCount(1)
        # formatted_data = [
        #     rslt['# Trades'],
        #     str(round(rslt['Win Rate [%]'],2))+'%',
        #     rslt['Equity Final [$]'],
        #     str(round(rslt['Return [%]'],2))+'%',
        #     rslt['Start'].date(),
        #     rslt['End'].date(),
        #     rslt['Duration'].days,
        #     str(round(rslt['Buy & Hold Return [%]'],2))+'%'
        # ]
        # for col_index, cell_data in enumerate(formatted_data):
        #     item = QTableWidgetItem(str(cell_data))
        #     item.setFont(font)
        #     item.setTextAlignment(Qt.AlignCenter)
        #     self.tableWidget_2.setItem(0, col_index, item) 
        # self.tableWidget_3.setRowCount(len(self.parameter50_data[0]))
        # for i in range(len(self.parameter50_data[0])):
        #     formatted_data = [
        #                         self.parameter50_data[0][i].replace('.TW',''), 
        #                         self.parameter50_name[i]
        #                      ]
        #     for col_index, cell_data in enumerate(formatted_data):
        #         item = QTableWidgetItem(str(cell_data))
        #         item.setFont(font)
        #         item.setTextAlignment(Qt.AlignCenter)
        #         self.tableWidget_3.setItem(i, col_index, item)
    def load_trades_data(self): #加载交易明细画面 啊這個就跟上面的一樣只是他不是默認的 基本要改的話就去改parameter_data
        font = QFont("Arial", 7*(self.window_height/525))
        import test1
        self.update_html()
        df, rslt, output_path = test1.test1_main(self.parameter_data,self.member_id, 1)
        if df.empty:
            QMessageBox.information(None, "Warning", "此股票在此時間段內還未上市")
            self.viewer = HtmlViewer('./HTML/white.html')
            self.splitter.insertWidget(0, self.viewer)
            return
        trades = rslt['_trades'][['EntryTime', 'ExitTime', 'EntryPrice', 'ExitPrice', 'Size', 'PnL', 'ReturnPct']]
        if trades.empty:
            QMessageBox.information(None, "Warning", "在此時間段內未達成交易")
            self.viewer = HtmlViewer('./HTML/white.html')
            self.splitter.insertWidget(0, self.viewer)
            return
        # 0050成分股之交易成果介面
        if self.checkBox.isChecked():
            # 顯示 tab_4
            self.tabWidget.setTabVisible(self.index_of_tab_4, True)
            self.tableWidget_3.setSortingEnabled(False)
            rslt50 = test1.test1_main(self.parameter50_data,self.member_id, 1)
            self.tableWidget_3.setRowCount(len(self.parameter50_data[0]))
            data_to_write = []  # 用于存储每行的数据
            for i in range(len(rslt50)):
                if rslt50[i]['# Trades'] == 'NA':
                    formatted_data = [
                        self.parameter50_data[0][i].replace('.TW',''),
                        self.parameter50_name[i],
                        'NA','NA','NA','NA','NA','NA','NA','NA'
                    ]
                else:
                    formatted_data = [
                        self.parameter50_data[0][i].replace('.TW',''),
                        self.parameter50_name[i],
                        rslt50[i]['# Trades'],
                        round(rslt50[i]['Win Rate [%]'],2),
                        rslt50[i]['Equity Final [$]'],
                        round(rslt50[i]['Return [%]'],2),
                        rslt50[i]['Start'].date(),
                        rslt50[i]['End'].date(),
                        rslt50[i]['Duration'].days,
                        round(rslt50[i]['Buy & Hold Return [%]'],2)
                    ]
                for col_index, cell_data in enumerate(formatted_data):
                    if isinstance(cell_data, (int, float)):
                        item = NumericTableWidgetItem(str(cell_data))
                    else:
                        item = QTableWidgetItem(str(cell_data))
                    item.setFont(font)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget_3.setItem(i, col_index, item)
                # 将当前行的数据添加到写入 Excel 的列表
                data_to_write.append(formatted_data)

            # 将数据列表转换为 DataFrame
            dataframe = pd.DataFrame(data_to_write, columns=[
                'Ticker', 'Name', '# Trades', 'Win Rate [%]', 'Equity Final [$]', 
                'Return [%]', 'Start', 'End', 'Duration (days)', 'Buy & Hold Return [%]'
            ])
            # 将 DataFrame 写入 Excel
            dataframe.to_excel('parameter50_data_file.xlsx', index=False)
            # 允許表格排序
            self.tableWidget_3.setSortingEnabled(True)
            self.tableWidget_3.cellClicked.connect(self.handle_cell_click_3)  # 點擊事件
        else:
            # 隱藏 tab_4
            self.tabWidget.setTabVisible(self.index_of_tab_4, False)

        # 觀察清單成分股之交易成果介面
        if self.checkBox_2.isChecked():
            # 顯示 tab_5
            self.tabWidget.setTabVisible(self.index_of_tab_5, True)
            self.tableWidget_4.setSortingEnabled(False)
            myrslt = test1.test1_main(self.my_parameter_data,self.member_id, 1)
            self.tableWidget_4.setRowCount(len(self.my_parameter_data[0]))
            for i in range(len(myrslt)):
                if myrslt[i]['# Trades'] == 'NA':
                    formatted_data = [
                        self.my_parameter_data[0][i].replace('.TW',''),
                        'NA','NA','NA','NA','NA','NA','NA','NA'
                    ]
                else:
                    formatted_data = [
                        self.my_parameter_data[0][i].replace('.TW',''),
                        myrslt[i]['# Trades'],
                        round(myrslt[i]['Win Rate [%]'],2),
                        myrslt[i]['Equity Final [$]'],
                        round(myrslt[i]['Return [%]'],2),
                        myrslt[i]['Start'].date(),
                        myrslt[i]['End'].date(),
                        myrslt[i]['Duration'].days,
                        round(myrslt[i]['Buy & Hold Return [%]'],2)
                    ]
                # print(formatted_data)
                for col_index, cell_data in enumerate(formatted_data):
                    if isinstance(cell_data, (int, float)):
                        item = NumericTableWidgetItem(str(cell_data))
                    else:
                        item = QTableWidgetItem(str(cell_data))
                    item.setFont(font)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget_4.setItem(i, col_index, item)
            # 允許表格排序
            self.tableWidget_4.setSortingEnabled(True)
            self.tableWidget_4.cellClicked.connect(self.handle_cell_click_4)
        else:
            # 隱藏 tab_5
            self.tabWidget.setTabVisible(self.index_of_tab_5, False)
        
        # 交易明细介面
        self.viewer = HtmlViewer('./HTML/Strategy.html')
        self.splitter.insertWidget(0, self.viewer)
        self.tableWidget.setRowCount(len(trades))
        # 允許表格排序
        self.tableWidget.setSortingEnabled(True)
        for row_index, row_data in enumerate(trades.itertuples(index=False)):
            formatted_data = [
                row_data.EntryTime.date(),
                row_data.ExitTime.date(),
                round(row_data.EntryPrice),
                round(row_data.ExitPrice),
                round(row_data.Size),
                round(row_data.PnL),
                round(row_data.ReturnPct,2)
            ]
            for col_index, cell_data in enumerate(formatted_data):
                if isinstance(cell_data, (int, float)):
                    item = NumericTableWidgetItem(str(cell_data))
                else:
                    item = QTableWidgetItem(str(cell_data))
                item.setFont(font)
                item.setTextAlignment(Qt.AlignCenter)
                self.tableWidget.setItem(row_index, col_index, item)

        # 交易成果介面
        self.tableWidget_2.setRowCount(1)
        # 允許表格排序
        self.tableWidget_2.setSortingEnabled(True)
        formatted_data = [
            rslt['# Trades'],
            round(rslt['Win Rate [%]'],2),
            rslt['Equity Final [$]'],
            round(rslt['Return [%]'],2),
            rslt['Start'].date(),
            rslt['End'].date(),
            rslt['Duration'].days,
            round(rslt['Buy & Hold Return [%]'],2)
        ]
        for col_index, cell_data in enumerate(formatted_data):
            if isinstance(cell_data, (int, float)):
                item = NumericTableWidgetItem(str(cell_data))
            else:
                item = QTableWidgetItem(str(cell_data))
            item.setFont(font)
            item.setTextAlignment(Qt.AlignCenter)
            self.tableWidget_2.setItem(0, col_index, item) 
        self.tableWidget_2.cellClicked.connect(self.handle_cell_click_2)
        self.printHint(df, rslt, output_path)
        # self.textBrowser.setHtml("""
        #                             <h2 style="color:red;">警告：</h2>
        #                             <h2 style="color:red;">警告：</h2>
        #                         """)
        
    def printHint(self, df, rslt, output_path):
        time_delta = pd.Timedelta(rslt["Duration"])
        times=rslt["# Trades"]
        # print(f"在這個策略中，平均每周做{times/time_delta.days*7:.2f}次交易，每月做{times/time_delta.days*30.4:.2f}次交易，每年做{times/time_delta.days*365:.2f}次交易")
        self.textBrowser.setHtml(f"<h1>在這個策略中，平均每周做{times/time_delta.days*7:.2f}次交易，每月做{times/time_delta.days*30.4:.2f}次交易，每年做{times/time_delta.days*365:.2f}次交易</h1>")
        # 假設您的回測數據已儲存在 DataFrame 中
        # 找出虧損超過 15% 的交易紀錄
        trade_lose = []

        for i in range(1, len(rslt["_trades"]['ExitTime'])):
            # 檢查 ReturnPct 是否小於 -0.15
            if rslt["_trades"]["ReturnPct"][i] < -0.15:
                # 提取 EntryTime 和 ExitTime 的年份和月份
                entry_month = pd.to_datetime(rslt["_trades"]["EntryTime"][i]).to_period("M")
                exit_month = pd.to_datetime(rslt["_trades"]["ExitTime"][i]).to_period("M")
                # 將進場和出場月份範圍儲存到 trade_lose 列表中
                trade_lose.append((entry_month, exit_month, rslt["_trades"]["ReturnPct"][i]))
        for i in range(0, len(trade_lose)):
            # print(f"在{trade_lose[i][0]}到{trade_lose[i][1]}損失{abs(trade_lose[i][2]*100):.2f}%")
            self.textBrowser.append(f"""<h2 style="color:red;">在{trade_lose[i][0]}到{trade_lose[i][1]}損失{abs(trade_lose[i][2]*100):.2f}%</h2">""")

        # 找高低點並計算斜率的函式，min_days:回推最小天數,max_days:回推最大天數,slope_threshold:斜率值(每天漲幅)
        min_days, max_days, slope_threshold = 15, 60, 0.0067
        df['Rolling_Max'] = df['Close'].rolling(window=10, center=True).max()
        high_points = df[df['Close'] == df['Rolling_Max']]
        slope_data = []
        for high_date, high_row in high_points.iterrows():
            high_price = high_row['Close']
            
            # 回推的時間範圍
            start_date = high_date - pd.Timedelta(days=max_days)
            end_date = high_date - pd.Timedelta(days=min_days)
            # 找出最低點
            range_data = df[(df.index >= start_date) & (df.index <= end_date)]
            
            if not range_data.empty:
                # 找到區間內的最低價和日期
                low_price = range_data['Close'].min()
                low_date = range_data['Close'].idxmin()
                # 漲幅
                increase_rate = (high_price - low_price) / low_price
                # 天數
                days_diff = (high_date - low_date).days
                # 計算斜率
                slope = increase_rate / days_diff if days_diff > 0 else None
                
                # 儲存符合條件的斜率
                if slope is not None and slope > slope_threshold:
                    slope_data.append({
                        'Start_Date': low_date, #Low_Date
                        'End_Date': high_date, #High_Date
                        'Low_Price': low_price,
                        'High_Price': high_price,
                        'Duration_Days': days_diff, #Days_Diff
                        'Increase_Rate': increase_rate,
                        #'Slope': slope #可註解
                    })
        # 將結果轉換為 DataFrame
        slope_df = pd.DataFrame(slope_data)
        if (not slope_df.empty):
            # 過濾重複的低點，依照days_diff排序後，保留相同low_date的最高值，在依照low_date排序
            slope_df = slope_df.sort_values('Duration_Days', ascending=False).drop_duplicates(subset=['Start_Date'], keep='first').sort_values('Start_Date')
            table_slope_html = slope_df.to_html(index=False, border=1, justify='center', escape=False)

            # 添加 HTML 样式和标题
            self.textBrowser.append(f"""<h2 style="color:red;">符合條件的大波段區間:""")
            self.textBrowser.append(f"""<div style="font-family: Arial, sans-serif; font-size: 14px;">{table_slope_html}</div>""")
        else:
            # 添加 HTML 样式和标题
            self.textBrowser.append(f"""<h2 style="color:red;">無符合條件的大波段區間""")

        n, k, min_consolidation_days=20, 2, 20
        df = df.copy()
        if not isinstance(df.index, pd.DatetimeIndex):
            df.index = pd.to_datetime(df.index)

        # 移動平均和標準差
        df['SMA'] = df['Close'].rolling(window=n).mean()
        df['StdDev'] = df['Close'].rolling(window=n).std()
        # 布林帶寬度
        df['Upper Band'] = df['SMA'] + k * df['StdDev']
        df['Lower Band'] = df['SMA'] - k * df['StdDev']
        df['Bollinger Band Width'] = df['Upper Band'] - df['Lower Band']

        # 計算布林帶寬度平均 window:回推天數
        df['Bollinger Band Width MA'] = df['Bollinger Band Width'].rolling(window=80).mean()

        # 判定是否為盤整期：布林帶寬度 < 寬度的75%
        df['Consolidation'] = df['Bollinger Band Width'] < df['Bollinger Band Width MA']*0.75

        # 找出每段連續盤整狀態的開始和結束日期，並確認持續天數
        consolidation_periods = []
        start_date = None
        consolidation_days = 0
        
        for date, row in df.iterrows():
            if row['Consolidation']:
                if start_date is None:
                    start_date = date  # 標記盤整期的開始
                consolidation_days += 1
            elif start_date is not None:
                # 結束盤整期
                if consolidation_days >= min_consolidation_days:
                    end_date = date
                    duration = (end_date - start_date).days
                    consolidation_periods.append([start_date, end_date, duration])
                start_date = None
                consolidation_days = 0

        result_df = pd.DataFrame(consolidation_periods, columns=['Start Date', 'End Date', 'Duration (days)'])
        if (not result_df.empty):
            # 转换 DataFrame 为 HTML 表格，去掉索引
            table_html = result_df.to_html(index=False, border=1, justify='center', escape=False)
            # 添加 HTML 样式和标题
            self.textBrowser.append(f"""<h2 style="color:red;">符合條件的盤整區間:</h2>""")
            self.textBrowser.append(f"""<div style="font-family: Arial, sans-serif; font-size: 14px;">{table_html}</div>""")
        else:
            # 添加 HTML 样式和标题
            self.textBrowser.append(f"""<h2 style="color:red;">無符合條件的盤整區間""")
        self.highlight_intervals_with_overlap(df, slope_data, consolidation_periods, output_path)
        self.highlight_trade_entries_and_exits(df, rslt['_trades'], output_path)


    def highlight_trade_entries_and_exits(self, df, trades, output_path):
        # 打開 Excel 檔案
        workbook = load_workbook(output_path)
        sheet = workbook["Stock Data"]

        # 設定顏色填充：進場點是綠色，出場點是紅色
        entry_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 綠色
        exit_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # 紅色

        # 遍歷交易數據並標註進場和出場點
        for _, trade in trades.iterrows():
            entry_time = pd.to_datetime(trade['EntryTime'])
            exit_time = pd.to_datetime(trade['ExitTime'])

            # 標註進場點
            for row in range(2, sheet.max_row + 1):
                cell_date = pd.Timestamp(sheet[f"A{row}"].value)  # 假設 'Date' 在列 A
                if cell_date == entry_time:
                    sheet[f"A{row}"].fill = entry_fill  # 只填充日期格
            # 標註出場點
            for row in range(2, sheet.max_row + 1):
                cell_date = pd.Timestamp(sheet[f"A{row}"].value)  # 假設 'Date' 在列 A
                if cell_date == exit_time:
                    sheet[f"A{row}"].fill = exit_fill  # 只填充日期格
        # 保存 Excel 檔案
        workbook.save(output_path)
        return

    def highlight_intervals_with_overlap(self, df, slope_data, consolidation_periods, output_path):
        # 將 DataFrame 匯出為 Excel
        # df.to_excel(output_path, index=False, sheet_name="Stock Data")  # 不匯出索引

        # 打開匯出的 Excel 檔案
        workbook = load_workbook(output_path)
        sheet = workbook["Stock Data"]

        # 定義顏色填充
        slope_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黃色
        consolidation_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 淺藍色
        overlap_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")  # 粉紅色（重疊部分）

        # 計算重疊區間
        overlap_periods = []
        for slope_interval in slope_data:
            for consolidation_interval in consolidation_periods:
                overlap_start = max(slope_interval['Start_Date'], consolidation_interval[0])  # 最大開始日期
                overlap_end = min(slope_interval['End_Date'], consolidation_interval[1])      # 最小結束日期
                if overlap_start <= overlap_end:  # 有交集
                    overlap_periods.append((overlap_start, overlap_end))

        # 高亮符合斜率條件的區間
        for interval in slope_data:
            start_date = interval['Start_Date']
            end_date = interval['End_Date']
            for row in range(2, sheet.max_row + 1):  # 第一行是標題，從第2行開始
                cell_date = pd.Timestamp(sheet[f"A{row}"].value)
                if start_date <= cell_date <= end_date:
                    for col in range(1, sheet.max_column + 1):  # 遍歷整行
                        sheet.cell(row=row, column=col).fill = slope_fill

        # 高亮符合盤整條件的區間
        for period in consolidation_periods:
            start_date, end_date, _ = period
            for row in range(2, sheet.max_row + 1):
                cell_date = pd.Timestamp(sheet[f"A{row}"].value)
                if start_date <= cell_date <= end_date:
                    for col in range(1, sheet.max_column + 1):  # 遍歷整行
                        sheet.cell(row=row, column=col).fill = consolidation_fill

        # 高亮重疊區間
        for overlap_start, overlap_end in overlap_periods:
            for row in range(2, sheet.max_row + 1):
                cell_date = pd.Timestamp(sheet[f"A{row}"].value)
                if overlap_start <= cell_date <= overlap_end:
                    for col in range(1, sheet.max_column + 1):  # 遍歷整行
                        sheet.cell(row=row, column=col).fill = overlap_fill

        # 保存 Excel 檔案
        workbook.save(output_path)
        return

    def handle_cell_click_2(self, row, column):
         # 尋找並移除之前的 HtmlViewer
        for i in range(self.splitter.count()):
            widget = self.splitter.widget(i)
            if isinstance(widget, HtmlViewer):  # 如果是 HtmlViewer，則移除
                widget.setParent(None)  # 移除部件的父級，這樣會自動釋放資源
                break  # 移除第一個找到的 HtmlViewer 後退出
        self.viewer = HtmlViewer('./HTML/Strategy.html')
        self.splitter.insertWidget(0, self.viewer)

    def handle_cell_click_3(self, row, column):
        item_id = self.tableWidget_3.item(row, 0).text()
        # 尋找並移除之前的 HtmlViewer
        for i in range(self.splitter.count()):
            widget = self.splitter.widget(i)
            if isinstance(widget, HtmlViewer):  # 如果是 HtmlViewer，則移除
                widget.setParent(None)  # 移除部件的父級，這樣會自動釋放資源
                break  # 移除第一個找到的 HtmlViewer 後退出
        # 插入新的 HtmlViewer
        if str(item_id)+'.TW' in self.parameter50_data[0]:
            self.viewer = HtmlViewer('./HTML/'+str(item_id)+'.html')
            self.splitter.insertWidget(0, self.viewer)
        else:
            self.viewer = HtmlViewer('./HTML/white.html')
            self.splitter.insertWidget(0, self.viewer)

    def handle_cell_click_4(self, row, column):
        item_id = self.tableWidget_4.item(row, 0).text()
        # 尋找並移除之前的 HtmlViewer
        for i in range(self.splitter.count()):
            widget = self.splitter.widget(i)
            if isinstance(widget, HtmlViewer):  # 如果是 HtmlViewer，則移除
                widget.setParent(None)  # 移除部件的父級，這樣會自動釋放資源
                break  # 移除第一個找到的 HtmlViewer 後退出
        # 插入新的 HtmlViewer
        if str(item_id)+'.TW' in self.my_parameter_data[0]:
            self.viewer = HtmlViewer('./HTML/'+str(item_id)+'.html')
            self.splitter.insertWidget(0, self.viewer)
        else:
            self.viewer = HtmlViewer('./HTML/white.html')
            self.splitter.insertWidget(0, self.viewer)

    def refresh_table_widget(self):
        self.table_widget.clearContents()
        self.table_widget.setRowCount(0)
        self.load_stock_data()

    def go_to_page_3(self):
        print("Switching to page 3")
        self.set_tableWidget_mylist_font()
        self.stackedWidget.setCurrentIndex(3)

    def toggle_table_widget_visibility(self):#table一开始的可见度
        if self.table_widget.isVisible():
            self.hide_table_with_animation()
        else:
            self.show_table_with_animation()

    def show_member_info(self):
        self.show_member()
    
    def on_cell_clicked(self, row, column):
        if column == 0:
            self.delete_row(row)
    
    def delete_row(self, row):
        self.table_widget.cellClicked.disconnect(self.on_cell_clicked)
        stock_symbol = self.table_widget.item(row, 1).text()
        self.remove_stock_from_db(stock_symbol)
        self.table_widget.removeRow(row)
        self.table_widget.cellClicked.connect(self.on_cell_clicked)
        self.my_parameter_data = [[],'KDCross','2014-07-26','2024-07-26','code','10000','0.002']
        if (self.check_target() == None):
            self.checkBox_2.setChecked(False)
            self.checkBox_2.setEnabled(False)
        else:
            self.checkBox_2.setEnabled(True)
            self.my_parameter_data[0] = []
            for id in self.check_target():
                self.my_parameter_data[0].append(id)

    def remove_stock_from_db(self, stock_symbol):
        cursor = self.db_connection.cursor()
        member_id = self.get_member_id()
        query = "DELETE FROM target WHERE target_id = %s AND member_id = %s"
        try:
            cursor.execute(query, (stock_symbol, member_id))
            self.db_connection.commit()
        except mysql.connector.Error as err:
            QMessageBox.warning(self, "Database Error", f"Error: {err}")
        cursor.close()

    def button_clicked_increase(self):
        print("轉移至新增策略")

    def on_header_clicked(self, logicalIndex):#不用管
        if logicalIndex == 0:
            text, ok = QInputDialog.getText(self, 'Input Dialog', '請輸入股票代號:')
            if ok:
                fetch_stock = FetchStock(self.get_favorite_symbols())
                stock_data = fetch_stock.fetch_stock_data(text)
                if "error" in stock_data:
                    QMessageBox.warning(self, "Error", f"Error fetching data for {text}")
                else:
                    if not self.is_stock_already_added(stock_data['symbol']):
                        self.update_table_with_stock_data(stock_data)
                        self.save_stock_data_to_db(stock_data)
                    else:
                        QMessageBox.warning(self, "Error", "這個股票代號已經在觀察清單中。")

    def on_header_clicked_mylist(self):#不用管
        self.load_public_list()

    def get_favorite_symbols(self):#不用管
        return ["2603", "4303"]

    def is_stock_already_added(self, stock_symbol):#不用管
        cursor = self.db_connection.cursor()
        member_id = self.get_member_id()
        query = "SELECT * FROM target WHERE target_id = %s AND member_id = %s"
        cursor.execute(query, (stock_symbol, member_id))
        result = cursor.fetchone()
        cursor.close()

    def update_table_with_stock_data(self, stock_data):#不用管
        new_data = [
            stock_data['symbol'],
            stock_data['close_price'],
            stock_data['price_change'],
            stock_data['price_change_percent']
        ]
        self.data.append(new_data)
        row_position = self.table_widget.rowCount()
        self.table_widget.insertRow(row_position)
        for column, cell_data in enumerate(new_data):
            item = QTableWidgetItem(cell_data)
            self.table_widget.setItem(row_position, column + 1, item)
            

        # 创建带图标的单元格
        item_with_icon = QTableWidgetItem()
        item_with_icon.setIcon(self.icon1)
        item_with_icon.setText("")
        self.table_widget.setItem(row_position, 0, item_with_icon)
        self.set_table_widget_font()
    def save_stock_data_to_db(self, stock_data):#不用管
        cursor = self.db_connection.cursor()
        member_id = self.get_member_id()
        query = "INSERT INTO target (target_id, member_id) VALUES (%s, %s)"
        try:
            cursor.execute(query, (stock_data['symbol'], member_id))
            self.db_connection.commit()
        except mysql.connector.Error as err:
            QMessageBox.warning(self, "Database Error", f"Error: {err}")
        cursor.close()
        self.my_parameter_data = [[],'KDCross','2014-07-26','2024-07-26','code','10000','0.002']
        if (self.check_target() == None):
            self.checkBox_2.setChecked(False)
            self.checkBox_2.setEnabled(False)
        else:
            self.checkBox_2.setEnabled(True)
            self.my_parameter_data[0] = []
            for id in self.check_target():
                self.my_parameter_data[0].append(id)

    def get_member_id(self):#不用管
        cursor = self.db_connection.cursor()
        query = "SELECT m_id FROM members WHERE name = %s"
        cursor.execute(query, (self.username,))
        result = cursor.fetchone()
        cursor.close()
        if result:
            return result[0]
        return None

    def load_stock_data(self):#不用管
        cursor = self.db_connection.cursor()
        member_id = self.get_member_id()
        query = "SELECT target_id FROM target WHERE member_id = %s"
        cursor.execute(query, (member_id,))
        stocks = cursor.fetchall()
        cursor.close()
        for stock in stocks:
            stock_symbol = stock[0]
            fetch_stock = FetchStock(self.get_favorite_symbols())
            stock_data = fetch_stock.fetch_stock_data(stock_symbol)
            if "error" not in stock_data:
                self.update_table_with_stock_data(stock_data)

    def start_timer(self):#不用管
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.refresh_stock_data)
        self.timer.start(5000)

    def refresh_stock_data(self):
        for i in range(self.table_widget.rowCount()):
            stock_symbol = self.table_widget.item(i, 1).text()
            fetch_stock = FetchStock(self.get_favorite_symbols())
            stock_data = fetch_stock.fetch_stock_data(stock_symbol)
            if "error" not in stock_data:
                # 创建 QTableWidgetItem 并设置内容
                close_price_item = QTableWidgetItem(stock_data['close_price'])
                price_change_item = QTableWidgetItem(stock_data['price_change'])
                price_change_percent_item = QTableWidgetItem(stock_data['price_change_percent'])


                # 设置 QTableWidgetItem 到表格中
                self.table_widget.setItem(i, 2, close_price_item)
                self.table_widget.setItem(i, 3, price_change_item)
                self.table_widget.setItem(i, 4, price_change_percent_item)
        self.set_table_widget_font()
    def show_table_with_animation(self):
        self.table_widget.setVisible(True)
    def hide_table_with_animation(self):
        self.table_widget.setVisible(False)
    def show_member(self):
        self.overlay = QWidget(self)
        self.overlay.setStyleSheet("background-color: rgba(0, 0, 0, 150);")
        self.overlay.setGeometry(self.rect())
        self.overlay.show()
        member_id = self.get_member_id()
        self.member_window = MemberInfoDialog(self.db_connection, member_id,self.window_width,self.window_height, self.big_label_font_size,self.small_label_font_size,self.ex_small_label_font_size,self)
        self.member_window.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.member_window.setAttribute(Qt.WA_TranslucentBackground)
        self.member_window.setWindowModality(Qt.ApplicationModal)
        self.member_window.resize(self.window_width/(962/511),self.window_height/(525/301))

        self.member_window.move(
            self.geometry().center() - self.member_window.rect().center()
        )

        self.opacity_effect = QGraphicsOpacityEffect(self.member_window)
        self.member_window.setGraphicsEffect(self.opacity_effect)

        self.fade_in_animation = QPropertyAnimation(self.opacity_effect, b"opacity")
        self.fade_in_animation.setDuration(500)
        self.fade_in_animation.setStartValue(0)
        self.fade_in_animation.setEndValue(1)
        self.fade_in_animation.start()

        self.member_window.show()

    def show_edit(self,row): #編輯畫面的設定
        self.overlay = QWidget(self)
        self.overlay.setStyleSheet("background-color: rgba(0, 0, 0, 150);")
        self.overlay.setGeometry(self.rect())
        self.overlay.show()
        self.edit_row = row
        self.edit_window = RoundedDialog(self)
        uic.loadUi('./圖片/edit_st.ui', self.edit_window)
        self.edit_window.setWindowModality(Qt.ApplicationModal)
        self.edit_window.resize(self.window_width/(962/481), self.window_height/(525/301))
        self.edit_window.move(
            self.geometry().center() - self.edit_window.rect().center()
        )

        # 处理按钮点击事件
        self.edit_window.pushButton_close.clicked.connect(self.close_edit_dialog)
        self.edit_window.pushButton_save.clicked.connect(self.save_strategy)
        #預設值
        strategy_name = self.tableWidget_mylist.item(self.edit_row, 0).text()
        description = self.tableWidget_mylist.item(self.edit_row, 1).text()

        self.edit_window.frame.setGeometry(self.window_width/(962/20),self.window_height/(525/10) , self.window_width/(962/451), self.window_height/(525/271))
        self.edit_window.label.setGeometry(self.window_width/(962/20),self.window_height/(525/20) , self.window_width/(962/51), self.window_height/(525/16))
        self.edit_window.label_2.setGeometry(self.window_width/(962/20),self.window_height/(525/50) , self.window_width/(962/31), self.window_height/(525/16))
        self.edit_window.label_3.setGeometry(self.window_width/(962/40),self.window_height/(525/237) , self.window_width/(962/31), self.window_height/(525/20))
        self.edit_window.public_true.setGeometry(self.window_width/(962/80),self.window_height/(525/240) , self.window_width/(962/31), self.window_height/(525/16))
        self.edit_window.public_false.setGeometry(self.window_width/(962/160),self.window_height/(525/240) , self.window_width/(962/31), self.window_height/(525/16))
        self.edit_window.pushButton_save.setGeometry(self.window_width/(962/300),self.window_height/(525/240) , self.window_width/(962/121), self.window_height/(525/31))
        self.edit_window.st_name.setGeometry(self.window_width/(962/90),self.window_height/(525/21) , self.window_width/(962/331), self.window_height/(525/20))
        self.edit_window.st_description.setGeometry(self.window_width/(962/20),self.window_height/(525/70) , self.window_width/(962/401), self.window_height/(525/31))
        self.edit_window.st_code.setGeometry(self.window_width/(962/20),self.window_height/(525/120) , self.window_width/(962/401), self.window_height/(525/111))
        self.edit_window.pushButton_save.setGeometry(self.window_width/(962/300),self.window_height/(525/240) , self.window_width/(962/121), self.window_height/(525/31))
        self.edit_window.label_15.setGeometry(self.window_width/(962/20),self.window_height/(525/99) , self.window_width/(962/71), self.window_height/(525/20))
        self.edit_window.pushButton_close.setGeometry(self.window_width/(962/430),0 , self.window_width/(962/20), self.window_height/(525/20))
        self.edit_window.pushButton_close.setIconSize(QSize(self.window_height/(525/41)/(41/10), self.window_height/(525/41)/(41/10))) 

        self.edit_window.label.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.edit_window.label_2.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.edit_window.label_15.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)

        self.edit_window.label_3.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.edit_window.public_true.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.edit_window.public_false.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.edit_window.pushButton_save.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.big_label_font_size}px;
                border: 1px solid black;
            }}
        """)
        self.edit_window.st_name.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.small_label_font_size}px;
             
            }}
        """)
        self.edit_window.st_description.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.small_label_font_size}px;
             
            }}
        """)
        self.edit_window.st_code.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.small_label_font_size}px;
             
            }}
        """)




        #連資料庫取得原本code內容
        edit_cursor = self.db_connection.cursor()
        m_id = self.get_member_id()

        query_code = """
        SELECT tempt_code
        FROM code 
        WHERE member_id = %s AND strategy = %s
        LIMIT 1
        """

        edit_cursor.execute(query_code,(m_id, strategy_name))
        result = edit_cursor.fetchone()
        code = result[0]
        edit_cursor.close()

        # 为弹窗中的输入框赋值
        self.edit_window.st_name.setPlainText(strategy_name)
        self.edit_window.st_description.setPlainText(description)
        self.edit_window.st_code.setPlainText(code)
        self.opacity_effect = QGraphicsOpacityEffect(self.edit_window)
        self.edit_window.setGraphicsEffect(self.opacity_effect)

        self.fade_in_animation = QPropertyAnimation(self.opacity_effect, b"opacity")
        self.fade_in_animation.setDuration(500)
        self.fade_in_animation.setStartValue(0)
        self.fade_in_animation.setEndValue(1)
        self.fade_in_animation.start()

        self.edit_window.show()

    def show_increase(self):#新增的設定
        self.overlay = QWidget(self)
        self.overlay.setStyleSheet("background-color: rgba(0, 0, 0, 150);")
        self.overlay.setGeometry(self.rect())
        self.overlay.show()

        self.increase_window = RoundedDialog(self)
        uic.loadUi('./圖片/edit_st.ui', self.increase_window)
        self.increase_window.setWindowModality(Qt.ApplicationModal)
        self.increase_window.resize(self.window_width/(962/481), self.window_height/(525/301))

        self.increase_window.move(
            self.geometry().center() - self.increase_window.rect().center()
        )
        self.increase_window.frame.setGeometry(self.window_width/(962/20),self.window_height/(525/10) , self.window_width/(962/451), self.window_height/(525/271))
        self.increase_window.label.setGeometry(self.window_width/(962/20),self.window_height/(525/20) , self.window_width/(962/51), self.window_height/(525/16))
        self.increase_window.label_2.setGeometry(self.window_width/(962/20),self.window_height/(525/50) , self.window_width/(962/31), self.window_height/(525/16))
        self.increase_window.label_3.setGeometry(self.window_width/(962/40),self.window_height/(525/237) , self.window_width/(962/31), self.window_height/(525/20))
        self.increase_window.public_true.setGeometry(self.window_width/(962/80),self.window_height/(525/240) , self.window_width/(962/31), self.window_height/(525/16))
        self.increase_window.public_false.setGeometry(self.window_width/(962/160),self.window_height/(525/240) , self.window_width/(962/31), self.window_height/(525/16))
        self.increase_window.pushButton_save.setGeometry(self.window_width/(962/300),self.window_height/(525/240) , self.window_width/(962/121), self.window_height/(525/31))
        self.increase_window.st_name.setGeometry(self.window_width/(962/90),self.window_height/(525/21) , self.window_width/(962/331), self.window_height/(525/20))
        self.increase_window.st_description.setGeometry(self.window_width/(962/20),self.window_height/(525/70) , self.window_width/(962/401), self.window_height/(525/31))
        self.increase_window.st_code.setGeometry(self.window_width/(962/20),self.window_height/(525/120) , self.window_width/(962/401), self.window_height/(525/111))
        self.increase_window.pushButton_save.setGeometry(self.window_width/(962/300),self.window_height/(525/240) , self.window_width/(962/121), self.window_height/(525/31))
        self.increase_window.label_15.setGeometry(self.window_width/(962/20),self.window_height/(525/99) , self.window_width/(962/71), self.window_height/(525/20))
        self.increase_window.pushButton_close.setGeometry(self.window_width/(962/430),0 , self.window_width/(962/20), self.window_height/(525/20))
        self.increase_window.pushButton_close.setIconSize(QSize(self.window_height/(525/41)/(41/10), self.window_height/(525/41)/(41/10))) 

        self.increase_window.label.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.increase_window.label_2.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.increase_window.label_15.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)

        self.increase_window.label_3.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.increase_window.public_true.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.increase_window.public_false.setStyleSheet(f"""
            font-size: {self.small_label_font_size}px;
        """)
        self.increase_window.pushButton_save.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.big_label_font_size}px;
                border: 1px solid black;
            }}
        """)
        self.increase_window.st_name.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.small_label_font_size}px;
             
            }}
        """)
        self.increase_window.st_description.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.small_label_font_size}px;
             
            }}
        """)
        self.increase_window.st_code.setStyleSheet(f"""
            QPushButton {{
                font-size: {self.small_label_font_size}px;
             
            }}
        """)
        # 处理按钮点击事件
        self.increase_window.pushButton_close.clicked.connect(self.close_increase_dialog)
        self.increase_window.pushButton_save.clicked.connect(self.increase_strategy)

        self.opacity_effect = QGraphicsOpacityEffect(self.increase_window)
        self.increase_window.setGraphicsEffect(self.opacity_effect)

        self.fade_in_animation = QPropertyAnimation(self.opacity_effect, b"opacity")
        self.fade_in_animation.setDuration(500)
        self.fade_in_animation.setStartValue(0)
        self.fade_in_animation.setEndValue(1)
        self.fade_in_animation.start()

        self.increase_window.show()
    def close_increase_dialog(self):
        if self.overlay:
            self.overlay.close()
        self.increase_window.close()
    def close_edit_dialog(self):
        if self.overlay:
            self.overlay.close()
        self.edit_window.close()
    def increase_strategy(self):#新增策略
        strategy = str(self.increase_window.st_name.toPlainText())
        description = str(self.increase_window.st_description.toPlainText())
        code = str(self.increase_window.st_code.toPlainText())

        # 检查哪个 QRadioButton 被选中并记录相应的值
        if self.increase_window.public_true.isChecked():
            status = '1'
        elif self.increase_window.public_false.isChecked():
            status = '0'
        else:
            status = '0'  # 默认状态

        try:
            # 创建数据库游标
            db_cursor = self.db_connection.cursor()

            # 检查是否已经存在相同 member_id 和 strategy 且 public = 0 的记录
            check_query_private = """
            SELECT COUNT(*) FROM code 
            WHERE member_id = %s AND strategy = %s AND public = 0
            """
            db_cursor.execute(check_query_private, (self.get_member_id(), strategy))
            result_private = db_cursor.fetchone()

            # 检查是否已经存在相同 strategy 且 public = 1 的记录
            check_query_public = """
            SELECT COUNT(*) FROM code 
            WHERE strategy = %s AND public = 1
            """
            db_cursor.execute(check_query_public, (strategy,))
            result_public = db_cursor.fetchone()

            if status == '1' and result_public[0] == 0:#result_private[0]
                # 如果 public = 1 并且不存在相同的记录，则插入新的记录
                insert_query = """
                INSERT INTO code (member_id, strategy, description, public, tempt_code, file_name) 
                VALUES (%s, %s, %s, %s, %s, %s)
                """
                db_cursor.execute(insert_query, (self.get_member_id(), strategy, description, status, code, (str(self.get_member_id()) + "_" + strategy)))
                insert_query = """
                INSERT INTO code (member_id, strategy, description, public, tempt_code, file_name) 
                VALUES (%s, %s, %s, %s, %s, %s)
                """
                db_cursor.execute(insert_query, (self.get_member_id(), strategy, description, 0, code, (str(self.get_member_id()) + "_" + strategy)))
                
                # 提交事务
                self.db_connection.commit()
                self.close_increase_dialog()  # 在保存成功后关闭弹窗和overlay
                self.fill_example_data()
                import test3
                test3.test3_main(self.get_member_id(), strategy)

            elif status == '0' and result_private[0] == 0: #result_public[0]
                # 如果 public = 0 并且不存在相同的 public = 1 的记录，则插入新的记录
                insert_query = """
                INSERT INTO code (member_id, strategy, description, public, tempt_code, file_name) 
                VALUES (%s, %s, %s, %s, %s, %s)
                """
                db_cursor.execute(insert_query, (self.get_member_id(), strategy, description, status, code, (str(self.get_member_id()) + "_" + strategy)))
                
                # 提交事务
                self.db_connection.commit()
                self.close_increase_dialog()  # 在保存成功后关闭弹窗和overlay
                self.fill_example_data()
                import test3
                test3.test3_main(self.get_member_id(), strategy)

            else:
                # 如果记录已存在，可以提示用户或选择其他操作
                QMessageBox.warning(self, "Duplicate Entry", "A strategy with the same name already exists for this user or as a public strategy.")

        except mysql.connector.Error as err:
            QMessageBox.warning(self, "Database Error", f"Error: {err}")
        
        finally:
            if db_cursor is not None:
                db_cursor.close()
    def save_strategy(self):#保存編輯策略
        strategy = str(self.edit_window.st_name.toPlainText())
        description = str(self.edit_window.st_description.toPlainText())
        code = str(self.edit_window.st_code.toPlainText())

        # 检查哪个 QRadioButton 被选中并记录相应的值
        if self.edit_window.public_true.isChecked():
            status = '1'
        elif self.edit_window.public_false.isChecked():
            status = '0'
        else:
            status = '0'  # 默认状态

        try:
            # 创建数据库游标
            db_cursor = self.db_connection.cursor()

            # 获取编辑前的策略名称
            original_strategy = self.tableWidget_mylist.item(self.edit_row, 0).text()

            # 如果用户选择将策略公开，先检查是否有相同名称的 public 策略
            if status == '1':
                check_query_public = """
                SELECT COUNT(*) FROM code 
                WHERE strategy = %s AND public = 1
                """
                db_cursor.execute(check_query_public, (strategy,))
                result_public = db_cursor.fetchone()
                #
                if result_public[0] > 0:
                    QMessageBox.warning(self, "Duplicate Entry", "A public strategy with the same name already exists.")
                    return

                # 插入新的公开策略，同时保留原本的私有策略
                insert_query = """
                INSERT INTO code (member_id, strategy, description, public, tempt_code, file_name) 
                VALUES (%s, %s, %s, %s, %s, %s)
                """
                db_cursor.execute(insert_query, (self.get_member_id(), strategy, description, status, code, (str(self.get_member_id()) + "_" + strategy)))
                import test3
                test3.test3_main(self.get_member_id(), strategy)
            else:
                # 如果策略名称没有更改，只更新描述和公开状态
                if strategy == original_strategy:
                    update_query = """
                    UPDATE code 
                    SET description = %s, public = %s ,tempt_code = %s 
                    WHERE member_id = %s AND strategy = %s
                    """
                    db_cursor.execute(update_query, (description, status, code, self.get_member_id(), strategy))
                    import test3
                    test3.test3_main(self.get_member_id(), strategy)
                else:
                    # 如果策略名称有更改，检查是否已存在相同的 private 策略
                    check_query_private = """
                    SELECT COUNT(*) FROM code 
                    WHERE member_id = %s AND strategy = %s AND public = 0
                    """
                    db_cursor.execute(check_query_private, (self.get_member_id(), strategy))
                    result_private = db_cursor.fetchone()

                    if result_private[0] > 0:
                        # 如果记录已存在，提示用户并退出保存操作
                        QMessageBox.warning(self, "Duplicate Entry", "你已經有這個策略了！")
                        return

                    # 更新原来的私有策略
                    update_query = """
                    UPDATE code 
                    SET strategy = %s, description = %s, public = %s, tempt_code = %s
                    WHERE member_id = %s AND strategy = %s AND public = 0
                    """
                    db_cursor.execute(update_query, (strategy, description, status, code, self.get_member_id(), original_strategy))
                    import test3
                    test3.test3_main(self.get_member_id(), strategy)


            # 提交事务
            self.db_connection.commit()
            self.close_edit_dialog()  # 在保存成功后关闭弹窗和overlay
            self.fill_example_data()

        except mysql.connector.Error as err:
            QMessageBox.warning(self, "Database Error", f"Error: {err}")
        
        finally:
            if db_cursor is not None:
                db_cursor.close()
    def closeEvent(self, event):
        if self.overlay:
            self.overlay.close()
        super().closeEvent(event)
    def go_to_page_1(self):
        self.stackedWidget.setCurrentIndex(1)

    def go_to_page_2(self):
        print("Switching to page 2")
        self.comboBox.clear()
        temp = self.find_my_strategy()
        if temp:  # 如果有结果
            for strategy in temp:
                self.comboBox.addItem(strategy[0])  # 将每个策略添加到 combobox
        # 刷新页面
        self.fill_example_data()
        self.load_public_list()
        self.stackedWidget.setCurrentIndex(2)
    def create_hover_event(self, button, scale_factor):
        def hover_event(event):
            if button == self.show_chatbot or self.public_bt:
                button.setIconSize(QSize(self.window_height/(525/41)/(41/25), self.window_height/(525/41)/(41/25)) * scale_factor)
            else:
                button.setIconSize(QSize(self.window_height/(525/41)/(41/20), self.window_height/(525/41)/(41/20)) * scale_factor)
        return hover_event

    def create_leave_event(self, button):
        def leave_event(event):
            if button == self.show_chatbot or self.public_bt:
                button.setIconSize(QSize(self.window_height/(525/41)/(41/25), self.window_height/(525/41)/(41/25)) )
            else:
                button.setIconSize(QSize(self.window_height/(525/41)/(41/20), self.window_height/(525/41)/(41/20)))
        return leave_event        
class MemberInfoDialog(QDialog):#会员资讯画面
    def __init__(self, db_connection, member_id,window_width,window_height,big_label_font,small_label_font,ex_small_label_font_size ,parent=None):
        super().__init__(parent)
        self.db_connection = db_connection
        self.member_id = member_id
        uic.loadUi("./圖片/member.ui", self)  # 加载 member.ui 文件
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        small_label_font+=2
        self.stackedWidget.setGeometry(-1,-1 , window_width/(962/511), window_height/(525/301))
        self.label_4.setGeometry(window_width/(962/50),window_height/(525/70) , window_width/(962/91), window_height/(525/20))
        self.label_9.setGeometry(window_width/(962/50),window_height/(525/150) , window_width/(962/61), window_height/(525/20))
        self.label_13.setGeometry(window_width/(962/50),window_height/(525/230) , window_width/(962/61), window_height/(525/20))

        self.pushButton.setGeometry(window_width/(962/250),window_height/(525/221) , window_width/(962/56), window_height/(525/20))
        self.pushButton_2.setGeometry(window_width/(962/460),0 , window_width/(962/20), window_height/(525/20))
        self.pushButton_2.setIconSize(QSize(window_height/(525/41)/(41/10), window_height/(525/41)/(41/10)))        
        self.stackedWidget.setGeometry(-1,-1, window_width/(962/511), window_height/(525/301))
        self.pushButton_3.setGeometry(window_width/(962/460),0 , window_width/(962/20), window_height/(525/20))
        self.pushButton_3.setIconSize(QSize(window_height/(525/41)/(41/10), window_height/(525/41)/(41/10)))
        self.pushButton_4.setGeometry(window_width/(962/30),window_height/(525/20) , window_width/(962/56), window_height/(525/17))
        self.pushButton_4.setIconSize(QSize(window_height/(525/41)/(41/20), window_height/(525/41)/(41/20)))
        
        self.label.setGeometry(window_width/(962/70),window_height/(525/66) , window_width/(962/41), window_height/(525/20))
        self.label_2.setGeometry(window_width/(962/70),window_height/(525/107) , window_width/(962/51), window_height/(525/20))
        self.label_3.setGeometry(window_width/(962/70),window_height/(525/162) , window_width/(962/71), window_height/(525/20))
        self.old_password.setGeometry(window_width/(962/160),window_height/(525/60) , window_width/(962/211), window_height/(525/20))
        self.new_password.setGeometry(window_width/(962/160),window_height/(525/100) , window_width/(962/211), window_height/(525/20))
        self.new_password_toggle.setGeometry(window_width/(962/145),window_height/(525/122) , window_width/(962/61), window_height/(525/20))
        self.confirm_password.setGeometry(window_width/(962/160),window_height/(525/160) , window_width/(962/211), window_height/(525/20))
        self.submit_button.setGeometry(window_width/(962/90),window_height/(525/230) , window_width/(962/231), window_height/(525/20))

        self.label_5.setGeometry(window_width/(962/20),window_height/(525/10) , window_width/(962/61), window_height/(525/20))
        self.label_6.setGeometry(window_width/(962/50),window_height/(525/50) , window_width/(962/71), window_height/(525/16))
        self.label_10.setGeometry(window_width/(962/50),window_height/(525/210) , window_width/(962/51), window_height/(525/20))
        self.label_name.setGeometry(window_width/(962/210),window_height/(525/61) , window_width/(962/231), window_height/(525/21))
        self.label_email.setGeometry(window_width/(962/210),window_height/(525/142) , window_width/(962/221), window_height/(525/21))
        
        self.pushButton.setStyleSheet(f"""
            font-size: {big_label_font}px;
            border:none;
        """)
        self.label_13.setStyleSheet(f"""
            font-size: {ex_small_label_font_size}px;
        """)
        self.label_9.setStyleSheet(f"""
            font-size: {ex_small_label_font_size}px;
        """)        
        self.label_4.setStyleSheet(f"""
            font-size: {ex_small_label_font_size}px;
        """)
        self.label_email.setStyleSheet(f"""
            font-size: {big_label_font}px;
        """)
        self.label_name.setStyleSheet(f"""
            font-size: {big_label_font}px;
        """)
        self.label_5.setStyleSheet(f"""
            font-size: {big_label_font}px;
        """)
        self.label_6.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.label_10.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.label.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.label_2.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.label_3.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.old_password.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.new_password.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.confirm_password.setStyleSheet(f"""
            font-size: {small_label_font}px;
        """)
        self.new_password_toggle.setStyleSheet(f"""
            QPushButton {{
                font-size: {small_label_font}px;
                border: none;
                color: #0072E3;
                padding: 5px 10px;
            }}
        """)
        
        self.submit_button.setStyleSheet(f"""
            QPushButton {{
                font-size: {small_label_font}px;
                background-color: lightblue;
                color: white;
                border-radius: 10px;
                padding: 5px 10px;
            }}
            QPushButton:hover {{
                background-color: #0072E3;
            }}
        """)
        self.label_name = self.findChild(QLabel, "label_name")  # 查找 QLabel 控件用于显示姓名
        self.label_email = self.findChild(QLabel, "label_email")  # 查找 QLabel 控件用于显示电子邮件
        self.pushButton = self.findChild(QPushButton, "pushButton")  # 查找 QPushButton 控件用于更改密码
        self.pushButton_2 = self.findChild(QPushButton, "pushButton_2")  # 查找 QPushButton 控件用于关闭对话框
        
        self.load_member_info()  # 加载会员信息

        self.pushButton.clicked.connect(self.change_password_page)  # 绑定更改密码按钮的点击事件
        self.pushButton_2.clicked.connect(self.close)  # 绑定关闭按钮的点击事件
        self.pushButton_3.clicked.connect(self.close)
        self.pushButton_4.clicked.connect(self.back_to_memberinfo)

        ##########################################################################
        self.old_password = self.findChild(QLineEdit, "old_password")
        self.new_password = self.findChild(QLineEdit, "new_password")
        self.new_password_toggle = self.findChild(QPushButton, "new_password_toggle")
        self.confirm_password = self.findChild(QLineEdit, "confirm_password")
        self.submit_button = self.findChild(QPushButton, "submit_button")

        # 设置旧密码输入框为密码模式
        self.old_password.setEchoMode(QLineEdit.Password)

        # 设置新密码输入框为密码模式
        self.new_password.setEchoMode(QLineEdit.Password)

        # 连接显示/隐藏密码按钮的点击事件
        self.new_password_toggle.clicked.connect(self.toggle_password_visibility)

        # 连接变更密码按钮的点击事件
        self.submit_button.clicked.connect(self.change_password)
        self.pushButton_2.clicked.connect(self.close)

    def load_member_info(self):
        cursor = self.db_connection.cursor()
        query = "SELECT name, email FROM members WHERE m_id = %s"
        cursor.execute(query, (self.member_id,))
        member = cursor.fetchone()
        if member:
            name, email = member
            self.label_name.setText(name)  # 将姓名设置到 QLabel
            self.label_email.setText(email)  # 将电子邮件设置到 QLabel
        else:
            QMessageBox.warning(self, "警告", "没有找到会员资料")
        cursor.close()

    def change_password_page(self):
        self.stackedWidget.setCurrentIndex(1)
    def back_to_memberinfo(self):
        self.stackedWidget.setCurrentIndex(0)


    def toggle_password_visibility(self):
        if self.new_password.echoMode() == QLineEdit.Password:
            self.new_password.setEchoMode(QLineEdit.Normal)
            self.new_password_toggle.setText("隱藏密碼")
        else:
            self.new_password.setEchoMode(QLineEdit.Password)
            self.new_password_toggle.setText("顯示密碼")

    def change_password(self):
        old_password = self.old_password.text()
        new_password = self.new_password.text()
        confirm_password = self.confirm_password.text()

        if new_password != confirm_password:
            QMessageBox.warning(self, "錯誤", "新密碼與確認密碼不一致")
            return

        hashed_old_password = hashlib.sha256(old_password.encode()).hexdigest()
        hashed_new_password = hashlib.sha256(new_password.encode()).hexdigest()

        cursor = self.db_connection.cursor()
        query = "SELECT password FROM members WHERE m_id = %s"
        cursor.execute(query, (self.member_id,))
        result = cursor.fetchone()
        if result and result[0] == hashed_old_password:
            update_query = "UPDATE members SET password = %s WHERE m_id = %s"
            cursor.execute(update_query, (hashed_new_password, self.member_id))
            self.db_connection.commit()
            QMessageBox.information(self, "成功", "密碼變更成功")
            self.close()
        else:
            QMessageBox.warning(self, "錯誤", "舊密碼不正確")
        cursor.close()
    def closeEvent(self, event):
        parent = self.parent()
        if parent and hasattr(parent, 'overlay') and parent.overlay:
            parent.overlay.close()
        super().closeEvent(event)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        rect = QRectF(self.rect().adjusted(0, 0, -1, -1))  # 调整绘制区域，使得矩形边界内缩1像素
        radius = 15
        path = QPainterPath()
        path.addRoundedRect(rect, radius, radius)
        painter.fillPath(path, QBrush(QColor(255, 255, 255, 255)))
        painter.setPen(Qt.NoPen)
        painter.drawPath(path)
class RoundedDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        rect = QRectF(self.rect().adjusted(0, 0, -1, -1))  # 调整绘制区域，使得矩形边界内缩1像素
        radius = 15
        path = QPainterPath()
        path.addRoundedRect(rect, radius, radius)
        painter.fillPath(path, QBrush(QColor(255, 255, 255, 255)))
        painter.setPen(Qt.NoPen)
        painter.drawPath(path)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    interface_window = LoginWindow()
    interface_window.show()
    sys.exit(app.exec_())
